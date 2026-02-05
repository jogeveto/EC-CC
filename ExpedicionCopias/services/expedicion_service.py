"""Servicio de orquestación para expedición de copias."""
import os
import re
import tempfile
import time
import requests
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Union, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from shared.utils.logger import get_logger
from ExpedicionCopias.core.crm_client import CRMClient, CasoNoEncontradoError
from ExpedicionCopias.core.docuware_client import DocuWareClient
from ExpedicionCopias.core.graph_client import GraphClient
from ExpedicionCopias.core.pdf_processor import PDFMerger
from ExpedicionCopias.core.file_organizer import FileOrganizer
from ExpedicionCopias.core.rules_engine import ExcepcionesValidator
from ExpedicionCopias.core.time_validator import TimeValidator
from ExpedicionCopias.core.non_critical_rules_validator import NonCriticalRulesValidator
from ExpedicionCopias.core.auth import Dynamics365Authenticator, AzureAuthenticator
from ExpedicionCopias.services.db_service import ExpedicionCopiasDB
from ExpedicionCopias.core.constants import (
    CAMPO_RADICADO_PRINCIPAL, CAMPO_RADICADO_OFICIALES,
    CAMPO_EMAIL_PARTICULARES, CAMPO_EMAIL_CREADOR,
    VALOR_DEFECTO_NA, RUTA_PARTICULARES, RUTA_OFICIALES, RUTA_OFICIALES_SLASH,
    PERMISO_READ, PERMISO_VIEW, ESTADO_EXITOSO, ESTADO_NO_EXITOSO, ESTADO_PENDIENTE,
    MENSAJE_PROCESADO_CORRECTAMENTE, MENSAJE_NO_PROCESADO,
    MSG_INTERRUMPIDO_FRANJA, MSG_NO_PROCESADO_FRANJA, MSG_NO_DOCUMENTOS,
    MSG_TODOS_EXCLUIDOS, MSG_TODAS_DESCARGAS_FALLARON, MSG_CASO_PROCESADO,
    VARIABLE_NOMBRE_SOCIEDAD, VARIABLE_NUMERO_PQRS, VARIABLE_FECHA_HOY,
    VARIABLE_CLIENTE, VARIABLE_CORREO_ELECTRONICO, VARIABLE_CORREO_ELECTRONICO_VARIANTE,
    VARIABLE_FECHA_INGRESO, VARIABLE_ENLACE_ONEDRIVE, VARIABLE_ENLACE_ONEDRIVE_VARIANTE,
    VARIABLE_FECHA_RESPUESTA, MSG_EMAIL_RESPONSABLE_NO_CONFIG, MSG_PLANTILLA_SIN_ADJUNTOS_NO_ENCONTRADA,
    MSG_PLANTILLA_REGLAS_NO_CRITICAS_NO_CONFIG, MSG_CUERPO_PLANTILLA_VACIO,
    MSG_ERROR_COMPARTIR, PLANTILLA_ERROR_COMPARTIR_SALUDO, PLANTILLA_ERROR_COMPARTIR_CUERPO,
    PLANTILLA_ERROR_COMPARTIR_CONTINUACION, PLANTILLA_ERROR_COMPARTIR_FIRMA,
    PLANTILLA_REGLAS_NO_CRITICAS_ASUNTO, MSG_FIRMA_NO_CONFIG, MSG_MODO_INVALIDO,
    MSG_EMAIL_QA_NO_CONFIG, HEADER_CODIGO_ASISTENTE, HEADER_CODIGO_BOT, HEADER_USUARIO_RED,
    HEADER_NOMBRE_ESTACION, HEADER_ID_PROCESO, HEADER_NO_RADICADO, HEADER_MATRICULAS,
    HEADER_ESTADO_PROCESO, HEADER_OBSERVACION, HEADER_FECHA_INICIO, HEADER_HORA_INICIO,
    HEADER_FECHA_FIN, HEADER_HORA_FIN, CODIGO_BOT_PARTICULARES, CODIGO_BOT_OFICIALES,
    VALOR_DEFECTO_CODIGO_ASISTENTE, VALOR_DEFECTO_USUARIO_RED, VALOR_DEFECTO_MAQUINA,
    RUTA_REPORTES, FORMATO_NOMBRE_REPORTE, MSG_PLANTILLA_EMAIL_NO_CONFIG,
    PLANTILLA_REPORTE_ASUNTO, PLANTILLA_REPORTE_CUERPO, TIPO_PARTICULARES, TIPO_OFICIALES,
    MSG_DATABASE_NO_CONFIG, MSG_DATABASE_PASSWORD_NO_CONFIG, MSG_ERROR_PROCESAMIENTO,
    PLANTILLA_ERROR_CASO_ASUNTO, PLANTILLA_ERROR_CASO_CUERPO, MESES_ESPAÑOL
)


class ExpedicionService:
    """Servicio principal para orquestar el proceso de expedición de copias."""

    def __init__(self, config: Dict[str, Any]) -> None:
        """
        Inicializa el servicio con configuración.

        Args:
            config: Diccionario con toda la configuración
        """
        self.config = config
        self.logger = get_logger("ExpedicionService")
        
        self.casos_procesados: List[Dict[str, Any]] = []
        self.casos_error: List[Dict[str, Any]] = []
        self.casos_pendientes: List[Dict[str, Any]] = []
        
        self._inicializar_clientes()
        self._inicializar_validadores()

    def _inicializar_clientes(self) -> None:
        """Inicializa los clientes de CRM, DocuWare y Graph."""
        dynamics_config = self.config.get("Dynamics365", {})
        graph_config = self.config.get("GraphAPI", {})
        
        dynamics_client_secret = self.config.get("dynamics_client_secret", "")
        graph_client_secret = self.config.get("graph_client_secret", "")
        
        if not dynamics_client_secret:
            raise ValueError("dynamics_client_secret no está configurado")
        if not graph_client_secret:
            raise ValueError("graph_client_secret no está configurado")
        
        dynamics_auth = Dynamics365Authenticator(
            tenant_id=dynamics_config.get("tenant_id", ""),
            client_id=dynamics_config.get("client_id", ""),
            client_secret=dynamics_client_secret
        )
        
        graph_auth = AzureAuthenticator(
            tenant_id=graph_config.get("tenant_id", ""),
            client_id=graph_config.get("client_id", ""),
            client_secret=graph_client_secret
        )
        
        self.crm_client = CRMClient(
            authenticator=dynamics_auth,
            base_url=dynamics_config.get("base_url", "")
        )
        
        self.graph_client = GraphClient(authenticator=graph_auth)
        
        # Las excepciones se cargarán desde la sección específica, inicializar con lista vacía
        rules_validator = ExcepcionesValidator([])
        
        self.docuware_client = DocuWareClient(
            config=self.config,
            rules_validator=rules_validator
        )

    def _inicializar_validadores(self) -> None:
        """Inicializa validadores de tiempo y reglas."""
        # Las franjas horarias y excepciones se cargarán desde la sección específica (Copias o CopiasOficiales)
        # Se inicializan con valores por defecto, se actualizarán en procesar_particulares/procesar_oficiales
        self.time_validator = TimeValidator(franjas_horarias=[])
        self.excepciones_validator = ExcepcionesValidator([])
        
        self.pdf_merger = PDFMerger()
        self.file_organizer = FileOrganizer()

    def _obtener_numero_radicado(self, caso: Dict[str, Any], fallback: str = VALOR_DEFECTO_NA) -> str:
        """
        Obtiene el número de radicado principal del caso.
        
        Args:
            caso: Diccionario con información del caso
            fallback: Valor a retornar si no se encuentra el radicado
            
        Returns:
            Número de radicado (sp_name) o fallback
        """
        return caso.get(CAMPO_RADICADO_PRINCIPAL, "") or fallback

    def _obtener_numero_radicado_oficiales(self, caso: Dict[str, Any], fallback: str = VALOR_DEFECTO_NA) -> str:
        """
        Obtiene el número de radicado específico para Oficiales.

        Este valor proviene del campo sp_nroderadicado y se usa únicamente
        para reemplazar el placeholder {numero_radicado} en las plantillas
        de email de CopiasOficiales, sin afectar el resto del proceso que
        sigue usando CAMPO_RADICADO_PRINCIPAL.
        """
        return caso.get(CAMPO_RADICADO_OFICIALES, "") or fallback

    def _validar_franja_horaria_tipo(self, tipo: str) -> bool:
        """
        Valida si el proceso debe ejecutarse según la franja horaria y día hábil.
        
        Args:
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
            
        Returns:
            True si debe ejecutarse, False en caso contrario
        """
        # Obtener configuración específica del tipo
        if tipo == "Copias":
            config_seccion = self.config.get("ReglasNegocio", {}).get("Copias", {})
        elif tipo == "CopiasOficiales":
            config_seccion = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
        else:
            self.logger.warning(f"Tipo desconocido: {tipo}. No se validará franja horaria.")
            return True  # Por defecto, permitir ejecución si el tipo es desconocido
        
        # Obtener franjas horarias de la configuración
        franjas_horarias = config_seccion.get("FranjasHorarias", [])
        
        # Crear TimeValidator temporal con las franjas horarias
        time_validator = TimeValidator(franjas_horarias=franjas_horarias)
        
        # Validar si debe ejecutarse
        return time_validator.debe_ejecutar()

    def _obtener_ruta_lock(self, tipo: str) -> Path:
        """
        Obtiene la ruta del archivo lock para un tipo de proceso.
        
        Args:
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
            
        Returns:
            Path del archivo lock
        """
        ruta_base = self.config.get("Globales", {}).get("RutaBaseProyecto", ".")
        ruta_base_path = Path(ruta_base)
        ruta_base_path.mkdir(parents=True, exist_ok=True)
        
        nombre_lock = f".lock_expedicion_{tipo.lower()}"
        return ruta_base_path / nombre_lock

    def _crear_lock(self, tipo: str) -> bool:
        """
        Crea un archivo lock para indicar que el proceso está ejecutándose.
        
        Args:
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
            
        Returns:
            True si se creó el lock, False si ya existe y está activo
        """
        lock_path = self._obtener_ruta_lock(tipo)
        
        # Verificar si el lock existe y está activo
        if lock_path.exists():
            try:
                # Leer timestamp del lock
                with open(lock_path, 'r') as f:
                    timestamp_str = f.read().strip()
                    timestamp = float(timestamp_str)
                    # Si el lock tiene menos de 24 horas, considerarlo activo
                    tiempo_transcurrido = datetime.now().timestamp() - timestamp
                    if tiempo_transcurrido < 86400:  # 24 horas en segundos
                        self.logger.warning(f"Lock existente encontrado para {tipo}. Proceso ya en ejecución.")
                        return False
                    else:
                        # Lock antiguo, eliminarlo
                        self.logger.warning(f"Lock antiguo encontrado para {tipo}. Eliminándolo.")
                        lock_path.unlink()
            except (ValueError, OSError) as e:
                self.logger.warning(f"Error leyendo lock existente: {e}. Eliminándolo.")
                lock_path.unlink()
        
        # Crear nuevo lock
        try:
            timestamp = datetime.now().timestamp()
            with open(lock_path, 'w') as f:
                f.write(str(timestamp))
            self.logger.info(f"Lock creado para {tipo} en {lock_path}")
            return True
        except Exception as e:
            self.logger.error(f"Error creando lock: {e}")
            return False

    def _eliminar_lock(self, tipo: str) -> None:
        """
        Elimina el archivo lock.
        
        Args:
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
        """
        lock_path = self._obtener_ruta_lock(tipo)
        try:
            if lock_path.exists():
                lock_path.unlink()
                self.logger.info(f"Lock eliminado para {tipo}")
        except Exception as e:
            self.logger.error(f"Error eliminando lock: {e}")

    def _validar_conexion_docuware(self) -> Tuple[bool, str]:
        """
        Valida la conexión a DocuWare intentando autenticar.
        
        Returns:
            Tupla con (exitoso: bool, mensaje_error: str)
        """
        try:
            self.logger.info("[VALIDACION] Iniciando validación de conexión a DocuWare...")
            self.docuware_client.autenticar()
            self.logger.info("[VALIDACION] Conexión a DocuWare validada exitosamente")
            return (True, "")
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"[VALIDACION] Error validando conexión a DocuWare: {error_msg}")
            return (False, error_msg)
        finally:
            # Cerrar sesión siempre después de la validación para liberar la licencia
            try:
                self.logger.info("[VALIDACION] Cerrando sesión de DocuWare después de validación...")
                self.docuware_client.cerrar_sesion()
                self.logger.info("[VALIDACION] Sesión de DocuWare cerrada exitosamente")
            except Exception as e:
                self.logger.warning(f"[VALIDACION] Error al cerrar sesión de DocuWare: {e}")

    def _validar_conexion_dynamics(self) -> Tuple[bool, str]:
        """
        Valida la conexión a Dynamics 365 intentando obtener el token (login).
        
        Returns:
            Tupla con (exitoso: bool, mensaje_error: str)
        """
        try:
            self.logger.info("[VALIDACION] Iniciando validación de conexión a Dynamics 365...")
            self.crm_client._get_token()
            self.logger.info("[VALIDACION] Conexión a Dynamics 365 validada exitosamente")
            return (True, "")
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"[VALIDACION] Error validando conexión a Dynamics 365: {error_msg}")
            return (False, error_msg)

    def _enviar_notificacion_error_conexion(self, tipo: str, servicio_fallido: str, mensaje_error: str) -> None:
        """
        Envía notificación de error de conexión al emailResponsable.
        
        Args:
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
            servicio_fallido: Nombre del servicio que falló ("DocuWare" o "Dynamics")
            mensaje_error: Mensaje de error detallado
        """
        try:
            # Obtener configuración según el tipo
            if tipo == "Copias":
                config_seccion = self.config.get("ReglasNegocio", {}).get("Copias", {})
                tipo_notificacion = "Copias"
            else:  # CopiasOficiales
                config_seccion = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
                tipo_notificacion = "CopiasOficiales"
            
            email_responsable = config_seccion.get("emailResponsable", "")
            
            if not email_responsable:
                self.logger.warning(f"emailResponsable no está configurado para {tipo}. No se enviará notificación de error de conexión.")
                return
            
            # Obtener plantilla de notificación
            notificaciones_config = self.config.get("Notificaciones", {})
            error_config = notificaciones_config.get("ErrorConexion", {})
            plantilla_config = error_config.get(tipo_notificacion, {})
            
            if not plantilla_config:
                self.logger.warning(f"Plantilla de notificación de error de conexión no encontrada para {tipo}. No se enviará notificación.")
                return
            
            asunto = plantilla_config.get("asunto", "")
            cuerpo = plantilla_config.get("cuerpo", "")
            
            if not asunto or not cuerpo:
                self.logger.warning(f"Plantilla de notificación de error de conexión incompleta para {tipo}. No se enviará notificación.")
                return
            
            # Reemplazar placeholders
            fecha_error = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cuerpo = cuerpo.replace("{fecha_error}", fecha_error)
            cuerpo = cuerpo.replace("{servicio_fallido}", servicio_fallido)
            cuerpo = cuerpo.replace("{mensaje_error}", mensaje_error)
            
            # Agregar firma al cuerpo
            cuerpo_con_firma = self._agregar_firma(cuerpo)
            
            # Enviar email
            destinatarios = self._obtener_destinatarios_por_modo(email_responsable)
            self.graph_client.enviar_email(
                usuario_id=self.config.get("GraphAPI", {}).get("user_email", ""),
                asunto=asunto,
                cuerpo=cuerpo_con_firma,
                destinatarios=destinatarios
            )
            self.logger.info(f"Notificación de error de conexión enviada exitosamente a {', '.join(destinatarios)} para {tipo}")
        except Exception as e:
            self.logger.error(f"Error enviando notificación de error de conexión para {tipo}: {e}")

    def _enviar_notificacion_inicio(self, tipo: str) -> None:
        """
        Envía notificación de inicio de ejecución al emailResponsable.
        
        Args:
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
        """
        try:
            # Obtener configuración según el tipo
            if tipo == "Copias":
                config_seccion = self.config.get("ReglasNegocio", {}).get("Copias", {})
                tipo_notificacion = "Copias"
            else:  # CopiasOficiales
                config_seccion = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
                tipo_notificacion = "CopiasOficiales"
            
            email_responsable = config_seccion.get("emailResponsable", "")
            
            if not email_responsable:
                self.logger.warning(f"emailResponsable no está configurado para {tipo}. No se enviará notificación de inicio.")
                return
            
            # Obtener plantilla de notificación
            notificaciones_config = self.config.get("Notificaciones", {})
            inicio_config = notificaciones_config.get("InicioEjecucion", {})
            plantilla_config = inicio_config.get(tipo_notificacion, {})
            
            if not plantilla_config:
                self.logger.warning(f"Plantilla de notificación de inicio no encontrada para {tipo}. No se enviará notificación.")
                return
            
            asunto = plantilla_config.get("asunto", "")
            cuerpo = plantilla_config.get("cuerpo", "")
            
            if not asunto or not cuerpo:
                self.logger.warning(f"Plantilla de notificación de inicio incompleta para {tipo}. No se enviará notificación.")
                return
            
            # Reemplazar placeholders
            fecha_inicio = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cuerpo = cuerpo.replace("{fecha_inicio}", fecha_inicio)
            
            # Agregar firma al cuerpo
            cuerpo_con_firma = self._agregar_firma(cuerpo)

            # Enviar email
            destinatarios = self._obtener_destinatarios_por_modo(email_responsable)
            self.graph_client.enviar_email(
                usuario_id=self.config.get("GraphAPI", {}).get("user_email", ""),
                asunto=asunto,
                cuerpo=cuerpo_con_firma,
                destinatarios=destinatarios
            )
            self.logger.info(f"Notificación de inicio enviada exitosamente a {', '.join(destinatarios)} para {tipo}")
        except Exception as e:
            self.logger.error(f"Error enviando notificación de inicio para {tipo}: {e}")

    def procesar_particulares(self) -> Dict[str, Any]:
        """
        Procesa casos de COPIAS (particulares).

        Returns:
            Diccionario con resultados del proceso
        """
        self.logger.info("[INICIO] Procesamiento de copias particulares")
        
        # Capturar timestamp de inicio
        fecha_hora_inicio = datetime.now()
        
        try:
            # Cargar configuración específica de Copias
            copias_config = self.config.get("ReglasNegocio", {}).get("Copias", {})
            
            # Actualizar validadores con configuración específica
            franjas_horarias = copias_config.get("FranjasHorarias", [])
            self.time_validator = TimeValidator(franjas_horarias=franjas_horarias)
            excepciones = copias_config.get("ExcepcionesDescarga", [])
            self.excepciones_validator = ExcepcionesValidator(excepciones)
            self.docuware_client.rules_validator = self.excepciones_validator
            # Actualizar configuración de carátula
            incluir_caratula = copias_config.get("incluirCaratula", False)
            self.docuware_client.incluir_caratula = incluir_caratula
            
            self.logger.info(f"Configuración cargada - Franjas horarias: {len(franjas_horarias)}, Excepciones: {len(excepciones)}")
            
            # La validación de franja horaria se hace antes de llamar a este método
            # No es necesario validar aquí, solo continuar con el procesamiento
            
            self.logger.info("Autenticando con DocuWare...")
            self.docuware_client.autenticar()
            self.logger.info("Autenticación con DocuWare exitosa")
            
            subcategorias = copias_config.get("Subcategorias", [])
            especificaciones = copias_config.get("Especificaciones", [])
            
            filtro = self._construir_filtro_crm(subcategorias, especificaciones)
            self.logger.info(f"Filtro CRM construido: {filtro}")
            
            casos = self.crm_client.consultar_casos(filtro)
            
            self.logger.info(f"Casos encontrados para procesar: {len(casos)}")
            
            # Inicializar validador de reglas no críticas
            non_critical_validator = NonCriticalRulesValidator(self.config)
            
            for caso in casos:
                # Validar franja horaria antes de procesar cada caso
                if not self.time_validator.debe_ejecutar():
                    case_id = caso.get('sp_documentoid', 'N/A')
                    radicado = self._obtener_numero_radicado(caso)
                    self.logger.warning(f"Fuera de franja horaria, interrumpiendo procesamiento. Caso {case_id} (Radicado: {radicado}) quedará pendiente.")
                    # Agregar caso actual y todos los restantes a pendientes
                    self.casos_pendientes.append({"caso": caso, "estado": "pendiente", "mensaje": MSG_INTERRUMPIDO_FRANJA})
                    # Agregar casos restantes a pendientes
                    indice_actual = casos.index(caso)
                    for caso_restante in casos[indice_actual + 1:]:
                        self.casos_pendientes.append({"caso": caso_restante, "estado": "pendiente", "mensaje": MSG_NO_PROCESADO_FRANJA})
                    break
                
                case_id = caso.get('sp_documentoid', 'N/A')
                radicado = self._obtener_numero_radicado(caso)
                matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
                matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
                self.logger.info(f"Procesando caso ID: {case_id}, Radicado: {radicado}, Matrículas: {', '.join(matriculas) if matriculas else 'N/A'}")
                
                # Validar reglas no críticas antes de procesar
                es_valido, mensaje_error = non_critical_validator.validar_reglas_no_criticas(caso, "Copias")
                if not es_valido:
                    self.logger.warning(f"[CASO {case_id}] Regla no crítica fallida: {mensaje_error}")
                    # Enviar email al responsable
                    self._enviar_email_regla_no_critica(caso, "Copias", mensaje_error)
                    # Agregar a casos_error con estado "No Exitoso"
                    self.casos_error.append({"caso": caso, "estado": "No Exitoso", "mensaje": mensaje_error})
                    # Continuar con el siguiente caso
                    continue
                
                try:
                    self._procesar_caso_particular(caso)
                    self.casos_procesados.append({"caso": caso, "estado": "exitoso"})
                    self.logger.info(f"Caso {case_id} procesado exitosamente")
                except Exception as e:
                    error_msg = str(e)
                    self.logger.error(f"Error procesando caso {case_id}: {error_msg}")
                    self.casos_error.append({"caso": caso, "estado": "error", "mensaje": error_msg})
                    self._manejar_error_caso(caso, error_msg)
            
            # Capturar timestamp de fin antes de generar reporte
            fecha_hora_fin = datetime.now()
            
            reporte_path = self._generar_reporte_excel("Copias", fecha_hora_inicio, fecha_hora_fin)
            self.logger.info(f"Reporte generado en: {reporte_path}")
            
            # Enviar reporte por email
            self._enviar_reporte_por_email("Copias", reporte_path, fecha_hora_inicio, fecha_hora_fin)
            
            return {
                "casos_procesados": len(self.casos_procesados),
                "casos_error": len(self.casos_error),
                "casos_pendientes": len(self.casos_pendientes),
                "reporte_path": reporte_path
            }
        finally:
            # Cerrar sesión siempre, incluso si hay errores
            try:
                self.logger.info("[FIN] Cerrando sesión de DocuWare...")
                self.docuware_client.cerrar_sesion()
                self.logger.info("[FIN] Sesión de DocuWare cerrada exitosamente")
            except Exception as e:
                self.logger.warning(f"[FIN] Error al cerrar sesión de DocuWare: {e}")

    def _procesar_caso_particular(self, caso: Dict[str, Any]) -> None:
        """Procesa un caso individual de particulares."""
        case_id = caso.get("sp_documentoid", "")
        matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
        matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
        
        self.logger.info(f"[CASO {case_id}] Iniciando procesamiento - Matrículas: {', '.join(matriculas) if matriculas else 'N/A'}")
        
        if not matriculas:
            raise ValueError("No se encontraron matrículas en el caso")
        
        ruta_temporal = tempfile.mkdtemp()
        self.logger.info(f"[CASO {case_id}] Ruta temporal creada: {ruta_temporal}")
        
        try:
            resultado_procesamiento = self._procesar_documentos_matricula(matriculas, ruta_temporal)
            documentos_descargados = resultado_procesamiento["documentos_descargados"]
            total_encontrados_docuware = resultado_procesamiento["total_encontrados_docuware"]
            total_disponibles = resultado_procesamiento["total_disponibles"]
            
            self.logger.info(f"[CASO {case_id}] Documentos descargados: {len(documentos_descargados)}")
            
            if not documentos_descargados:
                # Distinguir entre "no hay documentos" vs "todos excluidos"
                if total_encontrados_docuware == 0:
                    # No hay documentos en DocuWare - esto es un error
                    self.logger.error(f"[CASO {case_id}] No se encontraron documentos en DocuWare para las matrículas {', '.join(matriculas)}")
                    raise ValueError("No se encontraron documentos en DocuWare")
                elif total_disponibles == 0:
                    # Todos los documentos fueron excluidos por excepciones - esto NO es un error crítico pero debe aparecer en el reporte
                    self.logger.info(f"[CASO {case_id}] Todos los documentos fueron excluidos por excepciones. No se descargaron documentos (comportamiento esperado según reglas de negocio)")
                    # Enviar email al responsable
                    self._enviar_email_sin_adjuntos(caso, "Copias")
                    # Agregar a casos_error con estado "No Exitoso" para que aparezca en el reporte
                    observacion = "Todos los documentos fueron excluidos por excepciones de reglas de negocio"
                    raise ValueError(observacion)
                else:
                    # Hay documentos disponibles pero las descargas fallaron
                    self.logger.error(f"[CASO {case_id}] No se descargaron documentos aunque había {total_disponibles} disponible(s). Todas las descargas fallaron.")
                    raise ValueError("Todas las descargas fallaron")
            
            pdf_unificado = os.path.join(ruta_temporal, f"unificado_{case_id}.pdf")
            self.logger.info(f"[CASO {case_id}] Unificando {len(documentos_descargados)} documentos en PDF: {pdf_unificado}")
            self.pdf_merger.merge_pdfs(documentos_descargados, pdf_unificado)
            
            tamanio_mb = os.path.getsize(pdf_unificado) / (1024 * 1024)
            self.logger.info(f"[CASO {case_id}] PDF unificado creado - Tamaño: {tamanio_mb:.2f} MB")
            
            subcategoria_id = caso.get("_sp_subcategoriapqrs_value", "")
            email_destino, mensaje_error_email = self._validar_y_obtener_email_destino(caso, "Copias")

            # Validar email antes de continuar
            if mensaje_error_email:
                self.logger.warning(f"[CASO {case_id}] {mensaje_error_email}. Email obtenido: '{email_destino}'")
                # Enviar notificación al responsable
                self._enviar_email_regla_no_critica(caso, "Copias", f"El caso no tiene un email válido: {mensaje_error_email}")
                # Agregar a casos_error con estado "No Exitoso"
                self.casos_error.append({"caso": caso, "estado": "No Exitoso", "mensaje": mensaje_error_email})
                # Continuar con el siguiente caso
                return

            self.logger.info(f"[CASO {case_id}] Email destino: {email_destino}, Subcategoría: {subcategoria_id}")
            
            if tamanio_mb < 28:
                self.logger.info(f"[CASO {case_id}] PDF pequeño ({tamanio_mb:.2f} MB < 28 MB) - Enviando como adjunto por email")
                cuerpo_enviado = self._enviar_pdf_pequeno(pdf_unificado, subcategoria_id, email_destino, caso)
                self.logger.info(f"[CASO {case_id}] Email enviado exitosamente con adjunto")
            else:
                self.logger.info(f"[CASO {case_id}] PDF grande ({tamanio_mb:.2f} MB >= 28 MB) - Subiendo a OneDrive y enviando link")
                cuerpo_enviado = self._enviar_pdf_grande(pdf_unificado, case_id, subcategoria_id, email_destino, caso)
                self.logger.info(f"[CASO {case_id}] PDF subido a OneDrive y email con link enviado exitosamente")
            
            self.logger.info(f"[CASO {case_id}] Actualizando caso en CRM...")
            try:
                self.crm_client.actualizar_caso(case_id, {
                    "sp_descripciondelasolucion": cuerpo_enviado,
                    "sp_resolvercaso": True
                })
                self.logger.info(f"[CASO {case_id}] Caso actualizado en CRM exitosamente")
            except CasoNoEncontradoError as e:
                # Regla no crítica #4: Caso no encontrado al intentar actualizar
                error_msg = str(e)
                self.logger.warning(f"[CASO {case_id}] Regla no crítica fallida: {error_msg}")
                # Enviar email al responsable
                self._enviar_email_regla_no_critica(caso, "Copias", error_msg)
                # Lanzar excepción para que se maneje como error y se agregue a casos_error
                raise ValueError(f"Regla no crítica: {error_msg}")
            
            self._auditar_caso(case_id, "exitoso", "Caso procesado correctamente")
            
        finally:
            import shutil
            if os.path.exists(ruta_temporal):
                self.logger.info(f"[CASO {case_id}] Limpiando ruta temporal: {ruta_temporal}")
                shutil.rmtree(ruta_temporal, ignore_errors=True)

    def _procesar_documentos_matricula(
        self, matriculas: List[str], ruta_temporal: str
    ) -> Dict[str, Any]:
        """
        Procesa y descarga documentos para una lista de matrículas.

        Args:
            matriculas: Lista de matrículas
            ruta_temporal: Ruta temporal para descargas

        Returns:
            Diccionario con:
            - 'documentos_descargados': Lista de rutas de documentos descargados
            - 'total_encontrados_docuware': Total de documentos encontrados en DocuWare (antes del filtro)
            - 'total_disponibles': Total de documentos disponibles después del filtro
        """
        documentos_descargados = []
        total_matriculas = len(matriculas)
        total_encontrados_acumulado = 0
        total_disponibles_acumulado = 0
        
        self.logger.info(f"[DESCARGAS] Iniciando procesamiento de {total_matriculas} matrícula(s)")
        
        for idx, matricula in enumerate(matriculas, 1):
            self.logger.info(f"[DESCARGAS] Procesando matrícula {idx}/{total_matriculas}: {matricula}")
            
            resultado_busqueda = self.docuware_client.buscar_documentos(matricula)
            documentos = resultado_busqueda["documentos"]
            total_encontrados_docuware = resultado_busqueda["total_encontrados"]
            total_disponibles = resultado_busqueda["total_disponibles"]
            
            total_encontrados_acumulado += total_encontrados_docuware
            total_disponibles_acumulado += total_disponibles
            
            self.logger.info(f"[DESCARGAS] Matrícula {matricula}: {total_encontrados_docuware} documento(s) encontrado(s) en DocuWare, {total_disponibles} disponible(s) después de aplicar filtros")
            
            if total_encontrados_docuware == 0:
                self.logger.warning(f"[DESCARGAS] Matrícula {matricula}: No se encontraron documentos en DocuWare")
            elif total_disponibles == 0:
                self.logger.info(f"[DESCARGAS] Matrícula {matricula}: Todos los documentos fueron excluidos por excepciones (comportamiento esperado según reglas de negocio)")
            
            descargas_exitosas = 0
            descargas_fallidas = 0
            
            for doc in documentos:
                doc_id = doc.get("Id", "")
                if not doc_id:
                    self.logger.warning(f"[DESCARGAS] Matrícula {matricula}: Documento sin ID, omitiendo")
                    continue
                
                try:
                    self.logger.info(f"[DESCARGAS] Matrícula {matricula}: Descargando documento ID {doc_id}")
                    ruta_descarga = self.docuware_client.descargar_documento(doc_id, doc, ruta_temporal)
                    documentos_descargados.append(ruta_descarga)
                    descargas_exitosas += 1
                    self.logger.info(f"[DESCARGAS] Matrícula {matricula}: Documento {doc_id} descargado exitosamente en {ruta_descarga}")
                except Exception as e:
                    descargas_fallidas += 1
                    self.logger.warning(f"[DESCARGAS] Matrícula {matricula}: Error descargando documento {doc_id}: {e}")
            
            self.logger.info(f"[DESCARGAS] Matrícula {matricula}: Resumen - Encontrados en DocuWare: {total_encontrados_docuware}, Disponibles: {total_disponibles}, Exitosos: {descargas_exitosas}, Fallidos: {descargas_fallidas}")
        
        self.logger.info(f"[DESCARGAS] Resumen general - Total matrículas: {total_matriculas}, Total documentos descargados: {len(documentos_descargados)}")
        
        return {
            "documentos_descargados": documentos_descargados,
            "total_encontrados_docuware": total_encontrados_acumulado,
            "total_disponibles": total_disponibles_acumulado
        }

    def _enviar_pdf_pequeno(
        self, pdf_unificado: str, subcategoria_id: str, email_destino: str, caso: Dict[str, Any]
    ) -> str:
        """
        Envía PDF pequeño (< 28MB) como adjunto por email.

        Args:
            pdf_unificado: Ruta del PDF unificado
            subcategoria_id: ID de la subcategoría
            email_destino: Email del destinatario
            caso: Diccionario con información del caso

        Returns:
            Cuerpo del email enviado
        """
        plantilla = self._obtener_plantilla_email("Copias", subcategoria_id, "adjunto")
        
        # Reemplazar variables en asunto y cuerpo
        asunto_procesado = self._reemplazar_variables_plantilla(plantilla["asunto"], caso)
        cuerpo_procesado = self._reemplazar_variables_plantilla(plantilla["cuerpo"], caso)
        
        # Logging para verificar reemplazo
        case_id = caso.get("sp_documentoid", "N/A")
        self.logger.info(f"[CASO {case_id}] Asunto después de reemplazo: {asunto_procesado[:100]}...")
        self.logger.debug(f"[CASO {case_id}] Cuerpo después de reemplazo (primeros 200 chars): {cuerpo_procesado[:200]}...")
        
        # Agregar firma al cuerpo
        cuerpo_con_firma = self._agregar_firma(cuerpo_procesado)
        
        usuario_email = self.config.get("GraphAPI", {}).get("user_email", "")
        
        # Logging justo antes de enviar el email (estado final)
        self.logger.info(f"[CASO {case_id}] === ESTADO FINAL ANTES DE ENVIAR EMAIL ===")
        self.logger.info(f"[CASO {case_id}] Asunto final procesado: {asunto_procesado}")
        self.logger.info(f"[CASO {case_id}] Cuerpo final procesado (primeros 200 chars): {cuerpo_con_firma[:200]}...")
        self.logger.info(f"[CASO {case_id}] Destinatario: {email_destino}")
        self.logger.info(f"[CASO {case_id}] === FIN ESTADO FINAL ===")
        
        # Enviar el email
        self.graph_client.enviar_email(
            usuario_id=usuario_email,
            asunto=asunto_procesado,
            cuerpo=cuerpo_con_firma,
            destinatarios=self._obtener_destinatarios_por_modo([email_destino]),
            adjuntos=[pdf_unificado]
        )
        
        # Esperar un momento para que el email se guarde en Sent Items
        time.sleep(2)
        
        # Consultar el email enviado desde Office 365 y formatearlo
        try:
            email_enviado = self.graph_client.obtener_email_enviado(
                usuario_id=usuario_email,
                asunto=asunto_procesado
            )
            
            if email_enviado:
                email_formateado = self.graph_client.formatear_email_legible(
                    email_data=email_enviado,
                    caso=caso
                )
                self.logger.info(f"[CASO {caso.get('sp_documentoid', 'N/A')}] Email consultado y formateado desde Office 365")
                return email_formateado
            else:
                self.logger.warning(
                    f"[CASO {caso.get('sp_documentoid', 'N/A')}] "
                    f"No se pudo consultar el email desde Office 365. Usando cuerpo original como fallback."
                )
        except Exception as e:
            self.logger.warning(
                f"[CASO {caso.get('sp_documentoid', 'N/A')}] "
                f"Error consultando email desde Office 365: {str(e)}. Usando cuerpo original como fallback."
            )
        
        # Fallback: retornar el cuerpo original si no se pudo consultar el email
        return cuerpo_con_firma

    def _enviar_pdf_grande(
        self, pdf_unificado: str, case_id: str, subcategoria_id: str, email_destino: str, caso: Dict[str, Any]
    ) -> str:
        """
        Envía PDF grande (>= 28MB) subiéndolo a OneDrive y compartiendo link.

        Args:
            pdf_unificado: Ruta del PDF unificado
            case_id: ID del caso
            subcategoria_id: ID de la subcategoría
            email_destino: Email del destinatario
            caso: Diccionario con información del caso

        Returns:
            Cuerpo del email enviado con link
        """
        carpeta_base = self.config.get("OneDrive", {}).get("carpetaBase", "/ExpedicionCopias")
        usuario_email = self.config.get("GraphAPI", {}).get("user_email", "")
        
        # Usar sp_name para el nombre de la carpeta visible al usuario, con fallback a case_id
        nombre_caso = caso.get("sp_name", case_id)
        
        resultado_subida = self.graph_client.subir_a_onedrive(
            ruta_local=pdf_unificado,
            carpeta_destino=f"{carpeta_base}/Particulares/{nombre_caso}",
            usuario_id=usuario_email
        )
        
        # Usar la información del resultado de la subida directamente
        # Si no está disponible, intentar obtenerla de otra manera
        item_id = resultado_subida.get("id", "") if resultado_subida else ""
        web_url = resultado_subida.get("webUrl", "") if resultado_subida else ""
        
        # Si no se obtuvo la información de la subida, intentar obtenerla consultando el archivo
        if not item_id or not web_url:
            self.logger.warning(f"[CASO {case_id}] Información incompleta del resultado de subida, consultando archivo...")
            info_item = self.graph_client._obtener_info_carpeta(
                f"{carpeta_base}/Particulares/{nombre_caso}/{Path(pdf_unificado).name}",
                usuario_email
            )
            if info_item is None:
                raise ValueError(f"No se pudo obtener información del archivo subido en OneDrive para el caso {case_id}")
            item_id = info_item.get("id", "") if not item_id else item_id
            web_url = info_item.get("webUrl", "") if not web_url else web_url
        
        # Obtener emailResponsable de la configuración para compartir
        copias_config = self.config.get("ReglasNegocio", {}).get("Copias", {})
        email_responsable = copias_config.get("emailResponsable", "")
        
        # Intentar compartir el archivo con el responsable por correo
        link = ""
        try:
            if email_responsable:
                self.logger.info(f"[CASO {case_id}] Compartiendo archivo en OneDrive con responsable: {email_responsable}")
                link_info = self.graph_client.compartir_con_usuario(
                    item_id=item_id,
                    usuario_id=usuario_email,
                    email_destinatario=email_responsable,
                    rol=PERMISO_READ
                )
                link = link_info.get("link", "")
                if link:
                    self.logger.info(f"[CASO {case_id}] Archivo compartido con responsable exitosamente. Invitación enviada por correo.")
                else:
                    # Si el link viene vacío, usar webUrl como fallback
                    if web_url:
                        link = web_url
                        self.logger.warning(f"[CASO {case_id}] Link compartido vacío, usando webUrl como fallback: {link[:50]}...")
                    else:
                        self.logger.error(f"[CASO {case_id}] No se obtuvo link ni webUrl después de compartir")
            else:
                self.logger.warning(f"[CASO {case_id}] emailResponsable no configurado. Usando método de compartir tradicional.")
                # Fallback al método anterior si no hay emailResponsable configurado
                link_info = self.graph_client.compartir_carpeta(item_id, usuario_email)
                link = link_info.get("link", "")
                if link:
                    self.logger.info(f"[CASO {case_id}] Archivo compartido exitosamente (método tradicional)")
        except Exception as e:
            error_msg = str(e)
            self.logger.warning(f"[CASO {case_id}] No se pudo compartir el archivo en OneDrive: {error_msg}")
            
            # Usar webUrl como fallback
            if web_url:
                link = web_url
                self.logger.info(f"[CASO {case_id}] Usando URL directa de OneDrive como fallback: {link[:50]}...")
            else:
                # Si no hay webUrl, esto es un error crítico
                raise ValueError(f"No se pudo obtener enlace del archivo en OneDrive. Error al compartir: {error_msg}")
            
            # Enviar email al responsable notificando el error
            self._enviar_email_error_compartir(caso, "Copias", error_msg, link)
        
        # Validación final: si aún no hay link, usar webUrl o lanzar error
        if not link:
            if web_url:
                link = web_url
                self.logger.warning(f"[CASO {case_id}] Link final vacío, usando webUrl: {link[:50]}...")
            else:
                raise ValueError("No se obtuvo enlace del archivo en OneDrive")
        
        plantilla = self._obtener_plantilla_email("Copias", subcategoria_id, "onedrive")
        
        # Primero reemplazar {link} con el enlace de OneDrive
        cuerpo_con_link = plantilla["cuerpo"].replace("{link}", link)
        
        # Luego reemplazar todas las variables de la plantilla
        asunto_procesado = self._reemplazar_variables_plantilla(plantilla["asunto"], caso, link)
        cuerpo_procesado = self._reemplazar_variables_plantilla(cuerpo_con_link, caso, link)
        
        # Logging para verificar reemplazo
        self.logger.info(f"[CASO {case_id}] Asunto después de reemplazo: {asunto_procesado[:100]}...")
        self.logger.debug(f"[CASO {case_id}] Cuerpo después de reemplazo (primeros 200 chars): {cuerpo_procesado[:200]}...")
        
        # Agregar firma al cuerpo
        cuerpo_con_firma = self._agregar_firma(cuerpo_procesado)
        
        # Logging justo antes de enviar el email (estado final)
        self.logger.info(f"[CASO {case_id}] === ESTADO FINAL ANTES DE ENVIAR EMAIL ===")
        self.logger.info(f"[CASO {case_id}] Asunto final procesado: {asunto_procesado}")
        self.logger.info(f"[CASO {case_id}] Cuerpo final procesado (primeros 200 chars): {cuerpo_con_firma[:200]}...")
        self.logger.info(f"[CASO {case_id}] Destinatario: {email_destino}")
        self.logger.info(f"[CASO {case_id}] Link OneDrive: {link[:100] if link else 'N/A'}...")
        self.logger.info(f"[CASO {case_id}] === FIN ESTADO FINAL ===")
        
        # Enviar el email
        self.graph_client.enviar_email(
            usuario_id=usuario_email,
            asunto=asunto_procesado,
            cuerpo=cuerpo_con_firma,
            destinatarios=self._obtener_destinatarios_por_modo([email_destino])
        )
        
        # Esperar un momento para que el email se guarde en Sent Items
        time.sleep(2)
        
        # Consultar el email enviado desde Office 365 y formatearlo
        try:
            email_enviado = self.graph_client.obtener_email_enviado(
                usuario_id=usuario_email,
                asunto=asunto_procesado
            )
            
            if email_enviado:
                email_formateado = self.graph_client.formatear_email_legible(
                    email_data=email_enviado,
                    caso=caso
                )
                self.logger.info(f"[CASO {case_id}] Email consultado y formateado desde Office 365")
                return email_formateado
            else:
                self.logger.warning(
                    f"[CASO {case_id}] "
                    f"No se pudo consultar el email desde Office 365. Usando cuerpo original como fallback."
                )
        except Exception as e:
            self.logger.warning(
                f"[CASO {case_id}] "
                f"Error consultando email desde Office 365: {str(e)}. Usando cuerpo original como fallback."
            )
        
        # Fallback: retornar el cuerpo original si no se pudo consultar el email
        return cuerpo_con_firma

    def _enviar_email_sin_adjuntos(
        self, caso: Dict[str, Any], tipo: str
    ) -> None:
        """
        Envía email al responsable cuando no hay adjuntos disponibles por regla de negocio.

        Args:
            caso: Diccionario con información del caso
            tipo: Tipo de caso ("Copias" o "CopiasOficiales")
        """
        case_id = caso.get("sp_documentoid", "N/A")

        # Radicado general para logging y para Copias
        radicado = self._obtener_numero_radicado(caso)
        # Radicado específico de Oficiales (solo para placeholder {numero_radicado})
        radicado_oficial = self._obtener_numero_radicado_oficiales(caso, radicado)
        matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
        matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
        matriculas_str_display = ", ".join(matriculas) if matriculas else "N/A"
        
        self.logger.info(
            f"[CASO {case_id}] Enviando email al responsable - Tipo: {tipo}, "
            f"Radicado general: {radicado}, Radicado Oficiales: {radicado_oficial}"
        )
        
        # Obtener emailResponsable de la configuración
        if tipo == "Copias":
            config_seccion = self.config.get("ReglasNegocio", {}).get("Copias", {})
            subcategoria_id = caso.get("_sp_subcategoriapqrs_value", "")
        else:  # CopiasOficiales
            config_seccion = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
            subcategoria_id = ""
        
        email_responsable = config_seccion.get("emailResponsable", "")
        
        if not email_responsable:
            self.logger.warning(f"[CASO {case_id}] emailResponsable no está configurado para {tipo}. No se enviará email.")
            return
        
        # Obtener plantilla
        plantilla = self._obtener_plantilla_email(tipo, subcategoria_id, "sinAdjuntos")
        
        if not plantilla.get("asunto") or not plantilla.get("cuerpo"):
            self.logger.warning(f"[CASO {case_id}] Plantilla sinAdjuntos no encontrada o vacía para {tipo}. No se enviará email.")
            return
        
        # Reemplazar placeholders en la plantilla
        # Usar radicado como nombre del caso
        nombre_caso = radicado

        cuerpo = plantilla["cuerpo"].replace("{case_id}", case_id)
        cuerpo = cuerpo.replace("{ticket_number}", radicado)
        cuerpo = cuerpo.replace("{matriculas}", matriculas_str_display)

        # Si es CopiasOficiales, reemplazar además {numero_radicado} con el valor específico de Oficiales
        if tipo == "CopiasOficiales":
            cuerpo = cuerpo.replace("{numero_radicado}", radicado_oficial)
            asunto = plantilla["asunto"].replace("{numero_radicado}", radicado_oficial)
        else:
            asunto = plantilla["asunto"]
        
        # Agregar firma al cuerpo
        cuerpo_con_firma = self._agregar_firma(cuerpo)
        
        # Usar _obtener_destinatarios_por_modo para mantener lógica QA/PROD
        destinatarios = self._obtener_destinatarios_por_modo(email_responsable)
        
        try:
            self.graph_client.enviar_email(
                usuario_id=self.config.get("GraphAPI", {}).get("user_email", ""),
                asunto=asunto,
                cuerpo=cuerpo_con_firma,
                destinatarios=destinatarios
            )
            self.logger.info(f"[CASO {case_id}] Email enviado exitosamente al responsable: {', '.join(destinatarios)}")
        except Exception as e:
            self.logger.error(f"[CASO {case_id}] Error enviando email al responsable: {e}")

    def _enviar_email_regla_no_critica(
        self, caso: Dict[str, Any], tipo: str, novedad: str
    ) -> None:
        """
        Envía email al responsable cuando una regla no crítica falla.

        Args:
            caso: Diccionario con información del caso
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
            novedad: Mensaje descriptivo de la novedad identificada
        """
        case_id = caso.get("sp_documentoid", "N/A")
        self.logger.info(f"[CASO {case_id}] Enviando email de regla no crítica al responsable - Tipo: {tipo}")
        
        # Obtener emailResponsable de la configuración
        if tipo == "Copias":
            config_seccion = self.config.get("ReglasNegocio", {}).get("Copias", {})
        else:  # CopiasOficiales
            config_seccion = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
        
        email_responsable = config_seccion.get("emailResponsable", "")
        
        if not email_responsable:
            self.logger.warning(f"[CASO {case_id}] emailResponsable no está configurado para {tipo}. No se enviará email.")
            return
        
        # Obtener plantilla de reglas no críticas
        plantilla_config = config_seccion.get("PlantillaEmailReglasNoCriticas", {})
        
        if not plantilla_config:
            self.logger.warning(MSG_PLANTILLA_REGLAS_NO_CRITICAS_NO_CONFIG.format(tipo=tipo))
            return
        
        asunto = plantilla_config.get("asunto", PLANTILLA_REGLAS_NO_CRITICAS_ASUNTO)
        cuerpo = plantilla_config.get("cuerpo", "")
        
        if not cuerpo:
            self.logger.warning(MSG_CUERPO_PLANTILLA_VACIO)
            return
        
        # Reemplazar placeholder [Novedad identificada] con el mensaje específico
        cuerpo = cuerpo.replace("[Novedad identificada]", novedad)
        
        # Agregar firma al cuerpo
        cuerpo_con_firma = self._agregar_firma(cuerpo)
        
        # Usar _obtener_destinatarios_por_modo para mantener lógica QA/PROD
        destinatarios = self._obtener_destinatarios_por_modo(email_responsable)
        
        try:
            self.graph_client.enviar_email(
                usuario_id=self.config.get("GraphAPI", {}).get("user_email", ""),
                asunto=asunto,
                cuerpo=cuerpo_con_firma,
                destinatarios=destinatarios
            )
            self.logger.info(f"[CASO {case_id}] Email de regla no crítica enviado exitosamente al responsable: {', '.join(destinatarios)}")
        except Exception as e:
            self.logger.error(f"[CASO {case_id}] Error enviando email de regla no crítica al responsable: {e}")

    def _enviar_email_error_compartir(
        self, caso: Dict[str, Any], tipo: str, error_msg: str, link_onedrive: str
    ) -> None:
        """
        Envía email al responsable cuando falla el compartir archivo/carpeta en OneDrive.

        Args:
            caso: Diccionario con información del caso
            tipo: Tipo de caso ("Copias" o "CopiasOficiales")
            error_msg: Mensaje de error al compartir
            link_onedrive: URL directa de OneDrive del archivo/carpeta
        """
        case_id = caso.get("sp_documentoid", "N/A")
        radicado = self._obtener_numero_radicado(caso)
        
        self.logger.info(f"[CASO {case_id}] Enviando email al responsable por error al compartir - Tipo: {tipo}, Radicado: {radicado}")
        
        # Obtener emailResponsable de la configuración
        if tipo == "Copias":
            config_seccion = self.config.get("ReglasNegocio", {}).get("Copias", {})
        else:  # CopiasOficiales
            config_seccion = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
        
        email_responsable = config_seccion.get("emailResponsable", "")
        
        if not email_responsable:
            self.logger.warning(f"[CASO {case_id}] emailResponsable no está configurado para {tipo}. No se enviará email.")
            return
        
        # Construir mensaje HTML
        # Usar sp_name para el identificador visible al usuario, con fallback a case_id
        nombre_caso = self._obtener_numero_radicado(caso, case_id)
        asunto = MSG_ERROR_COMPARTIR.format(ticket_number=radicado)
        cuerpo = f"""<html><body>
<p>{PLANTILLA_ERROR_COMPARTIR_SALUDO}</p>
<p>{PLANTILLA_ERROR_COMPARTIR_CUERPO.format(nombre_caso=nombre_caso, ticket_number=radicado)}</p>
<p><strong>Información del caso:</strong></p>
<ul>
<li>ID Caso: {case_id}</li>
<li>Radicado: {radicado}</li>
<li>Tipo: {tipo}</li>
</ul>
<p><strong>Error:</strong> {error_msg}</p>
<p><strong>Enlace de OneDrive:</strong> <a href="{link_onedrive}">{link_onedrive}</a></p>
<p><strong>Nota:</strong> {PLANTILLA_ERROR_COMPARTIR_CONTINUACION}</p>
<p>{PLANTILLA_ERROR_COMPARTIR_FIRMA}</p>
</body></html>"""
        
        # Agregar firma al cuerpo
        cuerpo_con_firma = self._agregar_firma(cuerpo)
        
        # Usar _obtener_destinatarios_por_modo para mantener lógica QA/PROD
        destinatarios = self._obtener_destinatarios_por_modo(email_responsable)
        
        try:
            self.graph_client.enviar_email(
                usuario_id=self.config.get("GraphAPI", {}).get("user_email", ""),
                asunto=asunto,
                cuerpo=cuerpo_con_firma,
                destinatarios=destinatarios
            )
            self.logger.info(f"[CASO {case_id}] Email enviado exitosamente al responsable: {', '.join(destinatarios)}")
        except Exception as e:
            self.logger.error(f"[CASO {case_id}] Error enviando email al responsable: {e}")

    def _manejar_error_caso(self, caso: Dict[str, Any], error_msg: str) -> None:
        """
        Maneja el error de un caso enviando email si es posible.

        Args:
            caso: Diccionario con información del caso
            error_msg: Mensaje de error
        """
        try:
            email_destino, mensaje_error_email = self._validar_y_obtener_email_destino(caso, "Copias")
            if email_destino and not mensaje_error_email:
                self._enviar_email_error_caso(email_destino, caso, error_msg)
            else:
                self.logger.warning(f"[CASO {caso.get('sp_documentoid', 'N/A')}] No se puede enviar email de error: {mensaje_error_email}")
        except Exception as email_error:
            self.logger.error(f"Error enviando email de error: {email_error}")

    def procesar_oficiales(self) -> Dict[str, Any]:
        """
        Procesa casos de COPIAS ENTIDADES OFICIALES.

        Returns:
            Diccionario con resultados del proceso
        """
        self.logger.info("[INICIO] Procesamiento de copias oficiales")
        
        # Capturar timestamp de inicio
        fecha_hora_inicio = datetime.now()
        
        try:
            # Cargar configuración específica de CopiasOficiales
            copias_oficiales_config = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
            
            # Actualizar validadores con configuración específica
            franjas_horarias = copias_oficiales_config.get("FranjasHorarias", [])
            self.time_validator = TimeValidator(franjas_horarias=franjas_horarias)
            excepciones = copias_oficiales_config.get("ExcepcionesDescarga", [])
            self.excepciones_validator = ExcepcionesValidator(excepciones)
            self.docuware_client.rules_validator = self.excepciones_validator
            # Actualizar configuración de carátula
            incluir_caratula = copias_oficiales_config.get("incluirCaratula", False)
            self.docuware_client.incluir_caratula = incluir_caratula
            
            self.logger.info(f"Configuración cargada - Franjas horarias: {len(franjas_horarias)}, Excepciones: {len(excepciones)}")
            
            # La validación de franja horaria se hace antes de llamar a este método
            # No es necesario validar aquí, solo continuar con el procesamiento
            
            self.logger.info("Autenticando con DocuWare...")
            self.docuware_client.autenticar()
            self.logger.info("Autenticación con DocuWare exitosa")
            
            subcategorias = copias_oficiales_config.get("Subcategorias", [])
            especificaciones = copias_oficiales_config.get("Especificaciones", [])
            
            filtro = self._construir_filtro_crm(subcategorias, especificaciones)
            self.logger.info(f"Filtro CRM construido: {filtro}")
            
            casos = self.crm_client.consultar_casos(filtro)
            
            self.logger.info(f"Casos encontrados para procesar: {len(casos)}")
            
            # Inicializar validador de reglas no críticas
            non_critical_validator = NonCriticalRulesValidator(self.config)
            
            for caso in casos:
                # Validar franja horaria antes de procesar cada caso
                if not self.time_validator.debe_ejecutar():
                    case_id = caso.get('sp_documentoid', 'N/A')
                    radicado = self._obtener_numero_radicado(caso)
                    self.logger.warning(f"Fuera de franja horaria, interrumpiendo procesamiento. Caso {case_id} (Radicado: {radicado}) quedará pendiente.")
                    # Agregar caso actual y todos los restantes a pendientes
                    self.casos_pendientes.append({"caso": caso, "estado": "pendiente", "mensaje": MSG_INTERRUMPIDO_FRANJA})
                    # Agregar casos restantes a pendientes
                    indice_actual = casos.index(caso)
                    for caso_restante in casos[indice_actual + 1:]:
                        self.casos_pendientes.append({"caso": caso_restante, "estado": "pendiente", "mensaje": MSG_NO_PROCESADO_FRANJA})
                    break
                
                case_id = caso.get('sp_documentoid', 'N/A')
                radicado = self._obtener_numero_radicado(caso)
                matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
                matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
                self.logger.info(f"Procesando caso ID: {case_id}, Radicado: {radicado}, Matrículas: {', '.join(matriculas) if matriculas else 'N/A'}")
                
                # Validar reglas no críticas antes de procesar
                es_valido, mensaje_error = non_critical_validator.validar_reglas_no_criticas(caso, "CopiasOficiales")
                if not es_valido:
                    self.logger.warning(f"[CASO {case_id}] Regla no crítica fallida: {mensaje_error}")
                    # Enviar email al responsable
                    self._enviar_email_regla_no_critica(caso, "CopiasOficiales", mensaje_error)
                    # Agregar a casos_error con estado "No Exitoso"
                    self.casos_error.append({"caso": caso, "estado": "No Exitoso", "mensaje": mensaje_error})
                    # Continuar con el siguiente caso
                    continue
                
                try:
                    self._procesar_caso_oficial(caso)
                    self.casos_procesados.append({"caso": caso, "estado": "exitoso"})
                    self.logger.info(f"Caso {case_id} procesado exitosamente")
                except Exception as e:
                    error_msg = str(e)
                    self.logger.error(f"Error procesando caso {case_id}: {error_msg}")
                    self.casos_error.append({"caso": caso, "estado": "error", "mensaje": error_msg})
                    self._manejar_error_caso_oficial(caso, error_msg)
            
            # Capturar timestamp de fin antes de generar reporte
            fecha_hora_fin = datetime.now()
            
            reporte_path = self._generar_reporte_excel("CopiasOficiales", fecha_hora_inicio, fecha_hora_fin)
            self.logger.info(f"Reporte generado en: {reporte_path}")
            
            # Enviar reporte por email
            self._enviar_reporte_por_email("CopiasOficiales", reporte_path, fecha_hora_inicio, fecha_hora_fin)
            
            return {
                "casos_procesados": len(self.casos_procesados),
                "casos_error": len(self.casos_error),
                "casos_pendientes": len(self.casos_pendientes),
                "reporte_path": reporte_path
            }
        finally:
            # Cerrar sesión siempre, incluso si hay errores
            try:
                self.logger.info("[FIN] Cerrando sesión de DocuWare...")
                self.docuware_client.cerrar_sesion()
                self.logger.info("[FIN] Sesión de DocuWare cerrada exitosamente")
            except Exception as e:
                self.logger.warning(f"[FIN] Error al cerrar sesión de DocuWare: {e}")

    def _procesar_caso_oficial(self, caso: Dict[str, Any]) -> None:
        """Procesa un caso individual de entidades oficiales."""
        case_id = caso.get("sp_documentoid", "")
        # Usar sp_name para el radicado visible al usuario, con fallback a case_id
        radicado = self._obtener_numero_radicado(caso, case_id)
        matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
        matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
        
        self.logger.info(f"[CASO {case_id}] Iniciando procesamiento - Radicado: {radicado}, Matrículas: {', '.join(matriculas) if matriculas else 'N/A'}")
        
        if not matriculas:
            raise ValueError("No se encontraron matrículas en el caso")
        
        ruta_temporal = tempfile.mkdtemp()
        self.logger.info(f"[CASO {case_id}] Ruta temporal creada: {ruta_temporal}")
        
        try:
            resultado_procesamiento = self._procesar_documentos_oficiales(matriculas, ruta_temporal)
            documentos_info = resultado_procesamiento["documentos_info"]
            total_encontrados_docuware = resultado_procesamiento["total_encontrados_docuware"]
            total_disponibles = resultado_procesamiento["total_disponibles"]
            
            self.logger.info(f"[CASO {case_id}] Documentos descargados: {len(documentos_info)}")
            
            if not documentos_info:
                # Distinguir entre "no hay documentos" vs "todos excluidos"
                if total_encontrados_docuware == 0:
                    # No hay documentos en DocuWare - esto es un error
                    self.logger.error(f"[CASO {case_id}] No se encontraron documentos en DocuWare para las matrículas {', '.join(matriculas)}")
                    raise ValueError("No se encontraron documentos en DocuWare")
                elif total_disponibles == 0:
                    # Todos los documentos fueron excluidos por excepciones - esto NO es un error crítico pero debe aparecer en el reporte
                    self.logger.info(f"[CASO {case_id}] Todos los documentos fueron excluidos por excepciones. No se descargaron documentos (comportamiento esperado según reglas de negocio)")
                    # Enviar email al responsable
                    self._enviar_email_sin_adjuntos(caso, "CopiasOficiales")
                    # Agregar a casos_error con estado "No Exitoso" para que aparezca en el reporte
                    observacion = "Todos los documentos fueron excluidos por excepciones de reglas de negocio"
                    raise ValueError(observacion)
                else:
                    # Hay documentos disponibles pero las descargas fallaron
                    self.logger.error(f"[CASO {case_id}] No se descargaron documentos aunque había {total_disponibles} disponible(s). Todas las descargas fallaron.")
                    raise ValueError("Todas las descargas fallaron")
            
            self.logger.info(f"[CASO {case_id}] Organizando archivos por matrícula...")
            estructura = self.file_organizer.organizar_archivos(
                archivos=documentos_info,
                radicado=radicado,
                matriculas=matriculas,
                ruta_base=ruta_temporal
            )
            
            carpeta_organizada = estructura["ruta_base"]
            self.logger.info(f"[CASO {case_id}] Archivos organizados en: {carpeta_organizada}")
            
            self.logger.info(f"[CASO {case_id}] Subiendo carpeta a OneDrive y enviando email...")
            cuerpo_con_link = self._subir_y_enviar_carpeta_oficial(carpeta_organizada, caso)
            self.logger.info(f"[CASO {case_id}] Carpeta subida a OneDrive y email con link enviado exitosamente")
            
            self.logger.info(f"[CASO {case_id}] Actualizando caso en CRM...")
            try:
                self.crm_client.actualizar_caso(case_id, {
                    "sp_descripciondelasolucion": cuerpo_con_link,
                    "sp_resolvercaso": True
                })
                self.logger.info(f"[CASO {case_id}] Caso actualizado en CRM exitosamente")
            except CasoNoEncontradoError as e:
                # Regla no crítica #4: Caso no encontrado al intentar actualizar
                error_msg = str(e)
                self.logger.warning(f"[CASO {case_id}] Regla no crítica fallida: {error_msg}")
                # Enviar email al responsable
                self._enviar_email_regla_no_critica(caso, "CopiasOficiales", error_msg)
                # Lanzar excepción para que se maneje como error y se agregue a casos_error
                raise ValueError(f"Regla no crítica: {error_msg}")
            
            self._auditar_caso(case_id, "exitoso", "Caso procesado correctamente")
            
        finally:
            import shutil
            if os.path.exists(ruta_temporal):
                self.logger.info(f"[CASO {case_id}] Limpiando ruta temporal: {ruta_temporal}")
                shutil.rmtree(ruta_temporal, ignore_errors=True)

    def _procesar_documentos_oficiales(
        self, matriculas: List[str], ruta_temporal: str
    ) -> Dict[str, Any]:
        """
        Procesa y descarga documentos para entidades oficiales.

        Args:
            matriculas: Lista de matrículas
            ruta_temporal: Ruta temporal para descargas

        Returns:
            Diccionario con:
            - 'documentos_info': Lista de diccionarios con información de documentos descargados
            - 'total_encontrados_docuware': Total de documentos encontrados en DocuWare (antes del filtro)
            - 'total_disponibles': Total de documentos disponibles después del filtro
        """
        documentos_info = []
        total_matriculas = len(matriculas)
        total_encontrados_acumulado = 0
        total_disponibles_acumulado = 0
        
        self.logger.info(f"[DESCARGAS] Iniciando procesamiento de {total_matriculas} matrícula(s)")
        
        for idx, matricula in enumerate(matriculas, 1):
            self.logger.info(f"[DESCARGAS] Procesando matrícula {idx}/{total_matriculas}: {matricula}")
            
            resultado_busqueda = self.docuware_client.buscar_documentos(matricula)
            documentos = resultado_busqueda["documentos"]
            total_encontrados_docuware = resultado_busqueda["total_encontrados"]
            total_disponibles = resultado_busqueda["total_disponibles"]
            
            total_encontrados_acumulado += total_encontrados_docuware
            total_disponibles_acumulado += total_disponibles
            
            self.logger.info(f"[DESCARGAS] Matrícula {matricula}: {total_encontrados_docuware} documento(s) encontrado(s) en DocuWare, {total_disponibles} disponible(s) después de aplicar filtros")
            
            if total_encontrados_docuware == 0:
                self.logger.warning(f"[DESCARGAS] Matrícula {matricula}: No se encontraron documentos en DocuWare")
            elif total_disponibles == 0:
                self.logger.info(f"[DESCARGAS] Matrícula {matricula}: Todos los documentos fueron excluidos por excepciones (comportamiento esperado según reglas de negocio)")
            
            descargas_exitosas = 0
            descargas_fallidas = 0
            
            for doc in documentos:
                doc_id = doc.get("Id", "")
                if not doc_id:
                    self.logger.warning(f"[DESCARGAS] Matrícula {matricula}: Documento sin ID, omitiendo")
                    continue
                
                try:
                    self.logger.info(f"[DESCARGAS] Matrícula {matricula}: Descargando documento ID {doc_id}")
                    ruta_descarga = self.docuware_client.descargar_documento(doc_id, doc, ruta_temporal)
                    tipo_doc = self._obtener_tipo_documento(doc)
                    documentos_info.append({
                        "ruta": ruta_descarga,
                        "tipoDocumento": tipo_doc,
                        "matricula": matricula,
                        "documento": doc
                    })
                    descargas_exitosas += 1
                    self.logger.info(f"[DESCARGAS] Matrícula {matricula}: Documento {doc_id} (tipo: {tipo_doc}) descargado exitosamente en {ruta_descarga}")
                except Exception as e:
                    descargas_fallidas += 1
                    self.logger.warning(f"[DESCARGAS] Matrícula {matricula}: Error descargando documento {doc_id}: {e}")
            
            self.logger.info(f"[DESCARGAS] Matrícula {matricula}: Resumen - Encontrados en DocuWare: {total_encontrados_docuware}, Disponibles: {total_disponibles}, Exitosos: {descargas_exitosas}, Fallidos: {descargas_fallidas}")
        
        self.logger.info(f"[DESCARGAS] Resumen general - Total matrículas: {total_matriculas}, Total documentos descargados: {len(documentos_info)}")
        
        return {
            "documentos_info": documentos_info,
            "total_encontrados_docuware": total_encontrados_acumulado,
            "total_disponibles": total_disponibles_acumulado
        }

    def _subir_y_enviar_carpeta_oficial(
        self, carpeta_organizada: str, caso: Dict[str, Any]
    ) -> str:
        """
        Sube carpeta a OneDrive y envía email con link.

        Args:
            carpeta_organizada: Ruta de la carpeta organizada
            caso: Diccionario con información del caso

        Returns:
            Cuerpo del email enviado con link
        """
        carpeta_base = self.config.get("OneDrive", {}).get("carpetaBase", "/ExpedicionCopias")
        usuario_email = self.config.get("GraphAPI", {}).get("user_email", "")
        
        carpeta_destino = f"{carpeta_base}{RUTA_OFICIALES}"
        self.logger.info(f"[ONEDRIVE] Iniciando subida de carpeta organizada: {carpeta_organizada} -> {carpeta_destino}")
        
        info_carpeta = self.graph_client.subir_carpeta_completa(
            ruta_carpeta_local=carpeta_organizada,
            carpeta_destino=carpeta_destino,
            usuario_id=usuario_email
        )
        
        self.logger.info(f"[ONEDRIVE] Carpeta subida exitosamente. ID: {info_carpeta.get('id', 'N/A')}")
        
        carpeta_id = info_carpeta.get("id", "")
        if not carpeta_id:
            raise ValueError("No se obtuvo ID de carpeta después de la subida")
        
        web_url = info_carpeta.get("webUrl", "")
        case_id = caso.get("sp_documentoid", "N/A")
        
        # Para COPIAS OFICIALES, siempre usar emailResponsable del config
        copias_oficiales_config = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
        email_responsable = copias_oficiales_config.get("emailResponsable", "")
        
        # En modo QA, usar emailQa de la configuración global
        globales_config = self.config.get("Globales", {})
        modo = globales_config.get("modo", "PROD")
        
        if modo.upper() == "QA":
            email_qa = globales_config.get("emailQa", "")
            if email_qa:
                email_responsable = email_qa
                self.logger.info(f"[CASO {case_id}] Modo QA: Usando emailQa ({email_qa}) para compartir")
            else:
                self.logger.warning(f"[CASO {case_id}] Modo QA pero emailQa no configurado. Usando emailResponsable del config")
        
        # Validar que emailResponsable esté configurado antes de compartir
        if not email_responsable or not self._validar_email(email_responsable):
            error_msg = f"[CASO {case_id}] emailResponsable no está configurado o no es válido para CopiasOficiales. Valor: '{email_responsable}'"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Intentar compartir la carpeta con emailResponsable
        link = ""
        try:
            self.logger.info(f"[ONEDRIVE] Compartiendo carpeta (ID: {carpeta_id}) con emailResponsable: {email_responsable}")
            link_info = self.graph_client.compartir_con_usuario(
                item_id=carpeta_id,
                usuario_id=usuario_email,
                email_destinatario=email_responsable,
                rol=PERMISO_READ
            )
            link = link_info.get("link", "")
            if link:
                self.logger.info(f"[ONEDRIVE] Carpeta compartida con emailResponsable exitosamente. Invitación enviada por correo.")
            else:
                # Si el link viene vacío, usar webUrl como fallback
                if web_url:
                    link = web_url
                    self.logger.warning(f"[CASO {case_id}] Link compartido vacío, usando webUrl como fallback: {link[:50]}...")
                else:
                    self.logger.error(f"[CASO {case_id}] No se obtuvo link ni webUrl después de compartir")
        except Exception as e:
            error_msg = str(e)
            self.logger.warning(f"[CASO {case_id}] No se pudo compartir la carpeta en OneDrive: {error_msg}")
            
            # Usar webUrl como fallback
            if web_url:
                link = web_url
                self.logger.info(f"[CASO {case_id}] Usando URL directa de OneDrive como fallback: {link[:50]}...")
            else:
                # Si no hay webUrl, esto es un error crítico
                raise ValueError(f"No se pudo obtener enlace de la carpeta en OneDrive. Error al compartir: {error_msg}")
            
            # Enviar email al responsable notificando el error
            self._enviar_email_error_compartir(caso, "CopiasOficiales", error_msg, link)
        
        # Validación final: si aún no hay link, usar webUrl o lanzar error
        if not link:
            if web_url:
                link = web_url
                self.logger.warning(f"[CASO {case_id}] Link final vacío, usando webUrl: {link[:50]}...")
            else:
                raise ValueError("No se obtuvo enlace de la carpeta en OneDrive")
        
        # Construir la ruta completa de OneDrive
        nombre_carpeta = Path(carpeta_organizada).name
        ruta_onedrive = f"{carpeta_base}{RUTA_OFICIALES_SLASH}{nombre_carpeta}"
        self.logger.info(f"[ONEDRIVE] Ruta completa en OneDrive: {ruta_onedrive}")
        
        plantilla = self._obtener_plantilla_email("CopiasOficiales")

        # Obtener No. Radicado específico de Oficiales para el placeholder {numero_radicado}
        radicado_oficial = self._obtener_numero_radicado_oficiales(caso, "")
        self.logger.info(f"[CASO {case_id}] No. radicado (Oficiales) para email: '{radicado_oficial}' (sp_nroderadicado)")
        
        # Primero reemplazar {link}, {onedrive_path} y {numero_radicado}
        asunto_con_radicado = plantilla["asunto"].replace("{numero_radicado}", radicado_oficial)
        cuerpo_con_link = plantilla["cuerpo"].replace("{link}", link)
        cuerpo_con_link = cuerpo_con_link.replace("{onedrive_path}", ruta_onedrive)
        cuerpo_con_link = cuerpo_con_link.replace("{numero_radicado}", radicado_oficial)
        
        # Luego reemplazar todas las variables estándar de la plantilla
        asunto_procesado = self._reemplazar_variables_plantilla(asunto_con_radicado, caso, link)
        cuerpo_procesado = self._reemplazar_variables_plantilla(cuerpo_con_link, caso, link)
        
        # Agregar firma al cuerpo
        cuerpo_con_firma = self._agregar_firma(cuerpo_procesado)
        
        # Obtener y validar email_destino
        email_destino, mensaje_error_email = self._validar_y_obtener_email_destino(caso, "CopiasOficiales")

        if mensaje_error_email:
            error_msg = f"El caso no tiene un email válido: {mensaje_error_email}"
            self.logger.warning(f"[CASO {case_id}] {error_msg}")
            # Enviar notificación al responsable
            self._enviar_email_regla_no_critica(caso, "CopiasOficiales", error_msg)
            # Lanzar excepción para que se maneje como error no crítico
            raise ValueError(error_msg)

        self.logger.info(f"[CASO {case_id}] Email destino validado: {email_destino}")
        # Enviar el email
        self.graph_client.enviar_email(
            usuario_id=usuario_email,
            asunto=asunto_procesado,
            cuerpo=cuerpo_con_firma,
            destinatarios=self._obtener_destinatarios_por_modo([email_destino])
        )
        self.logger.info(f"[CASO {case_id}] Email enviado exitosamente a: {email_destino}")
        
        # Esperar un momento para que el email se guarde en Sent Items
        time.sleep(2)
        
        # Consultar el email enviado desde Office 365 y formatearlo
        try:
            email_enviado = self.graph_client.obtener_email_enviado(
                usuario_id=usuario_email,
                asunto=asunto_procesado
            )
            
            if email_enviado:
                email_formateado = self.graph_client.formatear_email_legible(
                    email_data=email_enviado,
                    caso=caso
                )
                self.logger.info(f"[CASO {case_id}] Email consultado y formateado desde Office 365")
                return email_formateado
            else:
                self.logger.warning(
                    f"[CASO {case_id}] "
                    f"No se pudo consultar el email desde Office 365. Usando cuerpo original como fallback."
                )
        except Exception as e:
            self.logger.warning(
                f"[CASO {case_id}] "
                f"Error consultando email desde Office 365: {str(e)}. Usando cuerpo original como fallback."
            )
        
        # Fallback: retornar el cuerpo original si no se pudo consultar el email
        return cuerpo_con_firma

    def _manejar_error_caso_oficial(self, caso: Dict[str, Any], error_msg: str) -> None:
        """
        Maneja el error de un caso oficial enviando email si es posible.

        Args:
            caso: Diccionario con información del caso
            error_msg: Mensaje de error
        """
        try:
            email_destino = self._obtener_email_creador(caso)
            if email_destino:
                self._enviar_email_error_caso(email_destino, caso, error_msg)
        except Exception as email_error:
            self.logger.error(f"Error enviando email de error: {email_error}")

    def _construir_filtro_crm(self, subcategorias: List[Union[str, Dict[str, Any]]], especificaciones: List[str]) -> str:
        """Construye el filtro OData para consultar casos en CRM."""
        # Extraer IDs de subcategorias (pueden ser strings o objetos con 'id')
        subcat_ids = []
        for subcat in subcategorias:
            if isinstance(subcat, str):
                subcat_ids.append(subcat)
            elif isinstance(subcat, dict):
                subcat_id = subcat.get("id", "")
                if subcat_id:
                    subcat_ids.append(subcat_id)
        
        condiciones_subcat = " or ".join([
            f"_sp_subcategoriapqrs_value eq '{subcat_id}'" for subcat_id in subcat_ids
        ])
        filtro_subcat = f"({condiciones_subcat})" if subcat_ids else ""
        
        condiciones_espec = " or ".join([
            f"_invt_especificacion_value eq '{espec_id}'" for espec_id in especificaciones
        ])
        filtro_espec = f"({condiciones_espec})" if especificaciones else ""
        
        partes = []
        if filtro_subcat:
            partes.append(filtro_subcat)
        if filtro_espec:
            partes.append(filtro_espec)
        
        partes.append("sp_resolvercaso eq false")
        
        return " and ".join(partes)

    def _obtener_plantilla_email(
        self, tipo: str, subcategoria_id: str = "", variante: str = ""
    ) -> Dict[str, str]:
        """Obtiene la plantilla de email según el tipo y variante."""
        if tipo == "CopiasOficiales":
            copias_oficiales_config = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
            plantillas = copias_oficiales_config.get("PlantillasEmail", {})
            # Para sinAdjuntos, buscar directamente en PlantillasEmail
            if variante == "sinAdjuntos":
                return plantillas.get("sinAdjuntos", {"asunto": "", "cuerpo": ""})
            return plantillas.get("default", {"asunto": "", "cuerpo": ""})
        
        if tipo == "Copias":
            copias_config = self.config.get("ReglasNegocio", {}).get("Copias", {})
            subcategorias = copias_config.get("Subcategorias", [])
            # Buscar la subcategoría por ID en el array de objetos
            for subcat in subcategorias:
                if isinstance(subcat, dict) and subcat.get("id") == subcategoria_id:
                    plantillas = subcat.get("PlantillasEmail", {})
                    return plantillas.get(variante, {"asunto": "", "cuerpo": ""})
            return {"asunto": "", "cuerpo": ""}
        
        return {"asunto": "", "cuerpo": ""}

    def _formatear_fecha_hoy_extendida(self) -> str:
        """
        Formatea la fecha de hoy en formato 'dd de mm de YYYY' (ej: "05 de enero de 2026").
        
        Returns:
            Fecha formateada en español
        """
        hoy = datetime.now()
        dia = hoy.day
        mes = MESES_ESPAÑOL[hoy.month - 1]
        año = hoy.year
        return f"{dia:02d} de {mes} de {año}"

    def _formatear_fecha_hoy_corta(self) -> str:
        """
        Formatea la fecha de hoy en formato 'dd/mm/YYYY' (ej: "05/01/2026").
        
        Returns:
            Fecha formateada
        """
        hoy = datetime.now()
        return hoy.strftime("%d/%m/%Y")

    def _formatear_fecha_createdon(self, createdon: Optional[str]) -> str:
        """
        Extrae solo la fecha de createdon (formato ISO) sin la hora.
        
        Args:
            createdon: Fecha en formato ISO (ej: "2026-01-05T10:30:00Z")
            
        Returns:
            Fecha en formato 'dd/mm/YYYY' o cadena vacía si no está disponible
        """
        if not createdon:
            return ""
        
        try:
            # Intentar parsear diferentes formatos de fecha ISO
            if 'T' in createdon:
                fecha_str = createdon.split('T')[0]
            else:
                fecha_str = createdon.split(' ')[0]
            
            fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
            return fecha_obj.strftime("%d/%m/%Y")
        except (ValueError, AttributeError):
            self.logger.warning(f"No se pudo formatear la fecha createdon: {createdon}")
            return ""

    def _extraer_cliente_del_titulo(self, titulo: Optional[str]) -> Optional[str]:
        """
        Extrae el nombre del cliente del título del PQRS.
        
        Patrón esperado: "Caso {nombre_cliente} {fecha} {hora}"
        Ejemplos:
        - "Caso EDISON  RODRIGUEZ HERNANDEZ  9/01/26 2:45 p.m."
        - "Caso YEAB CAPITAL S.A.S  9/01/26 3:01 p.m."
        - "Caso MARTHA ESTRADA  9/01/26 2:36 p.m."
        
        Args:
            titulo: Campo sp_titulopqrs del PQRS
        
        Returns:
            Nombre del cliente extraído o None si no se puede extraer
        """
        if not titulo:
            self.logger.debug(f"_extraer_cliente_del_titulo: Título vacío o None")
            return None
        
        # Patrón: "Caso " seguido de texto hasta encontrar fecha (formato dd/mm/yy o dd/mm/yyyy)
        # El patrón captura todo después de "Caso " hasta encontrar un patrón de fecha
        patron = r"Caso\s+(.+?)\s+\d{1,2}/\d{1,2}/\d{2,4}"
        match = re.search(patron, titulo, re.IGNORECASE)
        
        if match:
            nombre_cliente = match.group(1).strip()
            # Limpiar espacios múltiples
            nombre_cliente = re.sub(r'\s+', ' ', nombre_cliente)
            self.logger.debug(f"_extraer_cliente_del_titulo: Cliente extraído '{nombre_cliente}' del título '{titulo}'")
            return nombre_cliente if nombre_cliente else None
        
        self.logger.debug(f"_extraer_cliente_del_titulo: No se encontró match en título '{titulo}' con patrón '{patron}'")
        return None

    def _extraer_variables_plantilla(self, plantilla: str) -> List[str]:
        """
        Extrae todas las variables entre corchetes de la plantilla usando expresiones regulares.
        
        Args:
            plantilla: Texto de la plantilla con variables entre corchetes
            
        Returns:
            Lista de variables encontradas (sin los corchetes)
        """
        # Patrón regex para encontrar todas las variables entre corchetes: [variable]
        patron = r'\[([^\]]+)\]'
        variables_encontradas = re.findall(patron, plantilla)
        # Eliminar duplicados manteniendo el orden
        return list(dict.fromkeys(variables_encontradas))

    def _obtener_valor_variable(self, nombre_variable: str, caso: Dict[str, Any], link_onedrive: str = "") -> tuple[str, str]:
        """
        Obtiene el valor de una variable usando la lógica de múltiples campos alternativos.
        Sigue el patrón de pruebas-azure-connections usando .get() con or.
        
        Args:
            nombre_variable: Nombre de la variable (sin corchetes)
            caso: Diccionario con información del caso PQRS
            link_onedrive: Enlace de OneDrive (opcional)
            
        Returns:
            Tupla (valor, campo_usado) donde campo_usado indica qué campo(s) se usaron
        """
        nombre_variable_lower = nombre_variable.lower()
        
        # [Nombre de la sociedad] - usar solo _extraer_cliente_del_titulo (sin fallback)
        if "nombre de la sociedad" in nombre_variable_lower:
            titulo = caso.get("sp_titulopqrs")
            case_id = caso.get("sp_documentoid", "N/A")
            
            # Logging para debugging
            self.logger.info(f"[CASO {case_id}] [Nombre de la sociedad] - sp_titulopqrs disponible: '{titulo}'")
            
            cliente_del_titulo = self._extraer_cliente_del_titulo(titulo)
            
            if cliente_del_titulo:
                self.logger.info(f"[CASO {case_id}] [Nombre de la sociedad] - Cliente extraído del título: '{cliente_del_titulo}'")
                return (cliente_del_titulo, "sp_titulopqrs (extraído)")
            
            # Si no se puede extraer del título, retornar vacío (sin fallback)
            self.logger.warning(f"[CASO {case_id}] [Nombre de la sociedad] - No se pudo extraer del título (título: '{titulo}'), retornando vacío")
            return ("", "sp_titulopqrs (extraído)")
        
        # [Número PQRS] - usar método existente
        if "número pqrs" in nombre_variable_lower or "numero pqrs" in nombre_variable_lower:
            valor = self._obtener_numero_radicado(caso, "")
            return (valor, "sp_name (via _obtener_numero_radicado)")
        
        # [Fecha hoy] - fecha calculada
        if "fecha hoy" in nombre_variable_lower:
            valor = self._formatear_fecha_hoy_extendida()
            return (valor, "fecha_calculada")
        
        # [CLIENTE] - usar contacto_expandido obtenido con la misma lógica del proyecto de referencia
        if nombre_variable == "CLIENTE":
            case_id = caso.get("sp_documentoid", "N/A")
            
            # Intentar obtener contacto_expandido del objeto caso si está disponible
            contacto_expandido = caso.get("contacto_expandido")
            
            # Si no está disponible, intentar obtenerlo del CRM
            if not contacto_expandido:
                tiene_contacto = bool(caso.get("_sp_contacto_value"))
                
                if tiene_contacto:
                    # Posibles nombres de entidad para contacto (misma lógica del proyecto de referencia)
                    posibles_entidades_contacto = ["sp_contacto", "contact"]
                    pqrs_id = caso.get("sp_documentoid")
                    
                    if pqrs_id:
                        # Intentar expansión individual del contacto
                        for entidad in posibles_entidades_contacto:
                            try:
                                params_individual = {
                                    "$expand": entidad,
                                    "$select": "*"
                                }
                                response_individual = self.crm_client.get(
                                    f"/{self.crm_client.ENTITY_NAME}({pqrs_id})",
                                    params=params_individual
                                )
                                if entidad in response_individual:
                                    contacto_expandido = response_individual[entidad]
                                    self.logger.info(f"[CASO {case_id}] [CLIENTE] - Contacto expandido usando {entidad}")
                                    break
                            except Exception as e:
                                self.logger.debug(f"[CASO {case_id}] [CLIENTE] - Error intentando expandir con {entidad}: {str(e)[:100]}")
                                continue
            
            # Extraer el nombre del contacto
            if contacto_expandido:
                if isinstance(contacto_expandido, dict):
                    nombre_contacto = contacto_expandido.get("fullname") or contacto_expandido.get("sp_name") or contacto_expandido.get("name") or ""
                    if nombre_contacto:
                        self.logger.info(f"[CASO {case_id}] [CLIENTE] - Contacto encontrado: '{nombre_contacto}'")
                        return (nombre_contacto, "contacto_expandido")
            
            # Si no se encuentra, retornar vacío (sin fallback)
            self.logger.warning(f"[CASO {case_id}] [CLIENTE] - No se pudo obtener contacto expandido, retornando vacío")
            return ("", "contacto_expandido")
        
        # [Correo electrónico] o [Correo Electrónico] - buscar en múltiples campos
        if "correo electrónico" in nombre_variable_lower or "correo electronico" in nombre_variable_lower:
            valor = caso.get("invt_correoelectronico") or caso.get("sp_correoelectronico") or ""
            campo_usado = "invt_correoelectronico" if caso.get("invt_correoelectronico") else ("sp_correoelectronico" if caso.get("sp_correoelectronico") else "ninguno")
            return (valor, campo_usado)
        
        # [Fecha de ingreso de la solicitud] - fecha formateada
        if "fecha de ingreso" in nombre_variable_lower:
            createdon = caso.get("createdon", "")
            valor = self._formatear_fecha_createdon(createdon)
            return (valor, "createdon (formateado)")
        
        # [Enlace Onedrive.pdf] - enlace de OneDrive
        if "enlace onedrive" in nombre_variable_lower or "enlace onedrive.pdf" in nombre_variable_lower:
            valor = link_onedrive if link_onedrive else ""
            return (valor, "link_onedrive" if link_onedrive else "ninguno")
        
        # [Fecha de respuesta] - fecha calculada
        if "fecha de respuesta" in nombre_variable_lower:
            valor = self._formatear_fecha_hoy_corta()
            return (valor, "fecha_calculada")
        
        # Variable no reconocida
        return ("", "no_mapeada")

    def _reemplazar_variables_plantilla(
        self, plantilla: str, caso: Dict[str, Any], link_onedrive: str = ""
    ) -> str:
        """
        Reemplaza las variables entre corchetes en la plantilla según las reglas de oro.
        Extrae dinámicamente todas las variables y aplica la lógica de múltiples campos alternativos.
        
        Args:
            plantilla: Texto de la plantilla con variables entre corchetes
            caso: Diccionario con información del caso PQRS
            link_onedrive: Enlace de OneDrive (opcional)
            
        Returns:
            Plantilla con variables reemplazadas
        """
        case_id = caso.get("sp_documentoid", "N/A")
        resultado = plantilla
        
        # Extraer dinámicamente todas las variables entre corchetes
        variables_encontradas = self._extraer_variables_plantilla(plantilla)
        
        # Logging: Valores del caso antes del reemplazo
        self.logger.info(f"[CASO {case_id}] Valores del caso antes de reemplazar variables:")
        self.logger.info(f"[CASO {case_id}]   - sp_nombredelaempresa: '{caso.get('sp_nombredelaempresa', '')}'")
        self.logger.info(f"[CASO {case_id}]   - sp_titulopqrs: '{caso.get('sp_titulopqrs', '')}'")
        self.logger.info(f"[CASO {case_id}]   - invt_correoelectronico: '{caso.get('invt_correoelectronico', '')}'")
        self.logger.info(f"[CASO {case_id}]   - sp_correoelectronico: '{caso.get('sp_correoelectronico', '')}'")
        self.logger.info(f"[CASO {case_id}]   - sp_name: '{caso.get('sp_name', '')}'")
        
        # Logging: Variables encontradas en la plantilla
        if variables_encontradas:
            self.logger.info(f"[CASO {case_id}] Variables encontradas en la plantilla ({len(variables_encontradas)}): {', '.join(variables_encontradas)}")
        else:
            self.logger.info(f"[CASO {case_id}] No se encontraron variables entre corchetes en la plantilla")
        
        # Procesar cada variable encontrada
        variables_no_mapeadas = []
        for variable in variables_encontradas:
            # Obtener valor usando la lógica de múltiples campos alternativos
            valor, campo_usado = self._obtener_valor_variable(variable, caso, link_onedrive)
            
            # Logging detallado de cada variable
            if campo_usado == "no_mapeada":
                self.logger.warning(f"[CASO {case_id}] Variable [{variable}]: NO MAPEADA - no se encontró lógica de reemplazo")
                variables_no_mapeadas.append(variable)
            else:
                if valor:
                    self.logger.info(f"[CASO {case_id}] Variable [{variable}]: '{valor}' (campo usado: {campo_usado})")
                else:
                    self.logger.warning(f"[CASO {case_id}] Variable [{variable}]: VACÍA (campo usado: {campo_usado})")
            
            # Reemplazar la variable en la plantilla (con corchetes)
            variable_con_corchetes = f"[{variable}]"
            resultado = resultado.replace(variable_con_corchetes, valor)
        
        # Logging: Variables no mapeadas
        if variables_no_mapeadas:
            self.logger.warning(f"[CASO {case_id}] Variables que no se pudieron mapear: {', '.join(variables_no_mapeadas)}")
        
        # Logging: Plantilla final después del reemplazo
        self.logger.info(f"[CASO {case_id}] Plantilla después del reemplazo (primeros 200 chars): {resultado[:200]}...")
        
        return resultado

    def _agregar_firma(self, cuerpo: str) -> str:
        """
        Agrega la firma configurada al final del cuerpo del correo si no está presente.
        
        Args:
            cuerpo: Cuerpo del correo (HTML)
            
        Returns:
            Cuerpo del correo con la firma agregada (si no estaba presente)
        """
        firma_config = self.config.get("Firma", {})
        texto_firma = firma_config.get("texto", "")
        
        if not texto_firma:
            self.logger.warning(MSG_FIRMA_NO_CONFIG)
            return cuerpo
        
        # Verificar si la firma ya está presente en el cuerpo
        # Normalizar el texto de la firma para comparación (remover tags HTML y espacios)
        import re
        texto_firma_normalizado = re.sub(r'<[^>]+>', '', texto_firma).strip().lower()
        cuerpo_normalizado = re.sub(r'<[^>]+>', '', cuerpo).strip().lower()
        
        # Verificar si el texto normalizado de la firma ya está en el cuerpo
        if texto_firma_normalizado and texto_firma_normalizado in cuerpo_normalizado:
            self.logger.debug("La firma ya está presente en el cuerpo del correo. No se agregará nuevamente.")
            return cuerpo
        
        # Si el cuerpo termina con </body></html>, insertar la firma antes
        if cuerpo.rstrip().endswith("</body></html>"):
            # Remover el cierre de body y html
            cuerpo_sin_cierre = cuerpo.rstrip()[:-14]  # Remover </body></html>
            # Agregar firma y volver a cerrar
            return f"{cuerpo_sin_cierre}{texto_firma}</body></html>"
        else:
            # Si no tiene estructura HTML completa, simplemente agregar al final
            return f"{cuerpo}{texto_firma}"

    def _obtener_destinatarios_por_modo(self, destinatarios_originales: List[str]) -> List[str]:
        """
        Determina los destinatarios según el modo configurado (QA/PROD).
        
        Args:
            destinatarios_originales: Lista de emails destinatarios originales
            
        Returns:
            Lista de emails destinatarios (emailQa si modo es QA, originales si es PROD)
            
        Raises:
            ValueError: Si modo es QA y emailQa no está configurado
        """
        globales_config = self.config.get("Globales", {})
        modo = globales_config.get("modo", "PROD")
        
        # Validar modo
        modo_upper = modo.upper() if isinstance(modo, str) else "PROD"
        if modo_upper not in ["QA", "PROD"]:
            self.logger.warning(MSG_MODO_INVALIDO.format(modo=modo))
            modo_upper = "PROD"
        
        # Si es PROD, retornar destinatarios originales
        if modo_upper == "PROD":
            return destinatarios_originales
        
        # Si es QA, validar y usar emailQa
        email_qa = globales_config.get("emailQa", "")
        if not email_qa:
            error_msg = MSG_EMAIL_QA_NO_CONFIG
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Log cuando se redirige a QA
        destinatarios_str = ", ".join(destinatarios_originales) if destinatarios_originales else "N/A"
        self.logger.info(f"[MODO QA] Redirigiendo correo de destinatarios originales ({destinatarios_str}) a {email_qa}")
        
        return [email_qa]

    def _obtener_email_caso(self, caso: Dict[str, Any], tipo_proceso: str = "Copias") -> str:
        """
        Obtiene el email del caso usando únicamente invt_correoelectronico.
        
        Args:
            caso: Diccionario con información del caso
            tipo_proceso: Tipo de proceso (mantenido por compatibilidad, no afecta el resultado)
            
        Returns:
            Email del caso (invt_correoelectronico) o cadena vacía si no existe
        """
        # Siempre usar únicamente invt_correoelectronico, sin fallback
        return caso.get(CAMPO_EMAIL_CREADOR, "") or ""

    def _validar_y_obtener_email_destino(self, caso: Dict[str, Any], tipo: str) -> tuple[str, str]:
        """
        Valida y obtiene el email destino del caso.
        
        Args:
            caso: Diccionario con información del caso
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
            
        Returns:
            Tupla (email, mensaje_error):
            - email: Email válido o cadena vacía
            - mensaje_error: "Caso sin email" si está vacío, "Email mal formado" si es inválido, o "" si es válido
        """
        case_id = caso.get("sp_documentoid", "N/A")
        email = self._obtener_email_caso(caso, tipo)
        
        # Validar que no esté vacío
        if not email or not email.strip():
            return ("", "Caso sin email")
        
        # Validar formato con regex
        if not self._validar_email(email):
            return ("", "Email mal formado")
        
        return (email, "")

    def _validar_email(self, email: str) -> bool:
        """
        Valida si una cadena es un email válido.
        
        Args:
            email: Cadena a validar
            
        Returns:
            True si es un email válido, False en caso contrario
        """
        if not email or not isinstance(email, str):
            return False
        
        # Patrón básico de validación de email
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email.strip()))
    
    def _obtener_email_usuario_crm(self, user_id: str) -> str:
        """
        Consulta el email de un usuario en CRM usando su GUID.
        
        Args:
            user_id: GUID del usuario en CRM
            
        Returns:
            Email del usuario o cadena vacía si no se encuentra
        """
        if not user_id:
            self.logger.debug("[CRM] user_id vacío, no se puede consultar email")
            return ""
        
        # Validar que user_id no sea un email (para evitar consultas innecesarias)
        if self._validar_email(user_id):
            self.logger.debug(f"[CRM] user_id es un email válido, retornando directamente: {user_id}")
            return user_id
        
        try:
            # Consultar la entidad systemusers en Dynamics 365
            # Formato: /systemusers(guid) - OData acepta GUIDs con o sin llaves
            # Remover llaves si existen para usar formato estándar
            guid_formateado = user_id.strip().strip("{}")
            
            endpoint = f"/systemusers({guid_formateado})"
            params = {
                "$select": "internalemailaddress,domainname"
            }
            
            self.logger.debug(f"[CRM] Consultando email para usuario {user_id} en endpoint: {endpoint}")
            response = self.crm_client.get(endpoint, params=params)
            
            # Prioridad: internalemailaddress, luego domainname
            email = response.get("internalemailaddress", "") or response.get("domainname", "")
            
            if email and self._validar_email(email):
                self.logger.info(f"[CRM] Email obtenido para usuario {user_id}: {email}")
                return email
            else:
                self.logger.warning(f"[CRM] Email no válido o no encontrado para usuario {user_id}. Respuesta: {email}")
                return ""
                
        except requests.HTTPError as e:
            # Si el usuario no existe (404), no es un error crítico
            if e.response and e.response.status_code == 404:
                self.logger.warning(f"[CRM] Usuario {user_id} no encontrado en CRM (404)")
            else:
                self.logger.warning(f"[CRM] Error HTTP consultando email de usuario {user_id}: {e}")
            return ""
        except Exception as e:
            self.logger.warning(f"[CRM] Error consultando email de usuario {user_id}: {e}")
            return ""
    
    def _obtener_email_creador(self, caso: Dict[str, Any]) -> str:
        """
        Obtiene el email del creador/dueño del caso desde el CRM.
        
        Prioridad:
        1. sp_correoelectronico (email del destinatario del caso)
        2. Email del owner/creador consultado desde CRM
        
        Args:
            caso: Diccionario con información del caso
            
        Returns:
            Email válido o cadena vacía si no se encuentra
        """
        case_id = caso.get("sp_documentoid", "N/A")
        
        # Primera opción: usar sp_correoelectronico (email del destinatario del caso)
        email_destinatario = caso.get("sp_correoelectronico", "")
        if email_destinatario and self._validar_email(email_destinatario):
            self.logger.info(f"[CASO {case_id}] Usando sp_correoelectronico: {email_destinatario}")
            return email_destinatario
        
        # Segunda opción: consultar email del owner/creador desde CRM
        owner_id = caso.get("_ownerid_value", "") or caso.get("_createdby_value", "")
        if owner_id:
            email_usuario = self._obtener_email_usuario_crm(owner_id)
            if email_usuario:
                self.logger.info(f"[CASO {case_id}] Usando email del owner/creador: {email_usuario}")
                return email_usuario
            else:
                self.logger.warning(f"[CASO {case_id}] No se pudo obtener email del owner/creador (ID: {owner_id})")
        
        # Si no se encontró ningún email válido
        self.logger.warning(f"[CASO {case_id}] No se encontró email válido para el creador/dueño del caso")
        return ""

    def _obtener_tipo_documento(self, documento: Dict[str, Any]) -> str:
        """Obtiene el tipo de documento de los metadatos."""
        fields = documento.get("Fields", [])
        for field in fields:
            if field.get("FieldName") == "TRDNOMBREDOCUMENTO":
                item = field.get("Item")
                if item:
                    return str(item)
        return "SinTipo"

    def _enviar_email_error_caso(self, email: str, caso: Dict[str, Any], mensaje: str) -> None:
        """Envía email de error para un caso individual."""
        try:
            radicado = self._obtener_numero_radicado(caso)
            asunto = MSG_ERROR_PROCESAMIENTO.format(sp_ticketnumber=radicado)
            cuerpo = PLANTILLA_ERROR_CASO_CUERPO.format(mensaje=mensaje)
            
            # Agregar firma al cuerpo
            cuerpo_con_firma = self._agregar_firma(cuerpo)
            
            self.graph_client.enviar_email(
                usuario_id=self.config.get("GraphAPI", {}).get("user_email", ""),
                asunto=asunto,
                cuerpo=cuerpo_con_firma,
                destinatarios=self._obtener_destinatarios_por_modo([email])
            )
        except Exception as e:
            self.logger.error(f"Error enviando email de error: {e}")

    def _auditar_caso(self, case_id: str, estado: str, mensaje: str) -> None:
        """Stub para auditoría (no implementado por ahora)."""
        self.logger.info(f"Auditoría (stub): Caso {case_id} - {estado} - {mensaje}")

    def _generar_reporte_excel(self, tipo: str, fecha_hora_inicio: datetime, fecha_hora_fin: datetime) -> str:
        """
        Genera reporte Excel con los casos procesados.
        
        Args:
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
            fecha_hora_inicio: Timestamp de inicio de ejecución
            fecha_hora_fin: Timestamp de fin de ejecución
            
        Returns:
            Ruta del archivo Excel generado
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Expedición Copias"
        
        # Obtener datos del sistema con fallbacks
        reportes_config = self.config.get("Reportes", {})
        codigo_asistente = reportes_config.get("CodigoAsistente", "R_CCMA_ExpedicionCopias")
        
        # CodigoBot es dinámico según el tipo
        if tipo == "Copias":
            codigo_bot = "ExpedicionCopias_Particulares"
        else:  # CopiasOficiales
            codigo_bot = "ExpedicionCopias_Oficiales"
        
        # Usuario de red bot runner con fallback
        usuario_red = os.getenv('USERNAME', '')
        if not usuario_red:
            usuario_red = reportes_config.get("UsuaroRedBR", "usuario.red")
        
        # Nombre estación bot runner con fallback
        nombre_estacion = os.environ.get('COMPUTERNAME', '')
        if not nombre_estacion:
            nombre_estacion = reportes_config.get("NumeroMaquinaBR", "MAQUINA-001")
        
        # ID Proceso (PID del proceso Python)
        id_proceso = os.getpid()
        
        # Formatear fechas y horas
        fecha_inicio = fecha_hora_inicio.strftime("%Y-%m-%d")
        hora_inicio = fecha_hora_inicio.strftime("%H:%M:%S")
        fecha_fin = fecha_hora_fin.strftime("%Y-%m-%d")
        hora_fin = fecha_hora_fin.strftime("%H:%M:%S")
        
        # Definir headers con las nuevas columnas
        headers = [
            HEADER_CODIGO_ASISTENTE,
            HEADER_CODIGO_BOT,
            HEADER_USUARIO_RED,
            HEADER_NOMBRE_ESTACION,
            HEADER_ID_PROCESO,
            HEADER_NO_RADICADO,
            HEADER_MATRICULAS,
            HEADER_ESTADO_PROCESO,
            HEADER_OBSERVACION,
            HEADER_FECHA_INICIO,
            HEADER_HORA_INICIO,
            HEADER_FECHA_FIN,
            HEADER_HORA_FIN
        ]
        ws.append(headers)
        
        # Formatear headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar casos exitosos
        for item in self.casos_procesados:
            caso = item.get("caso", {})
            radicado = self._obtener_numero_radicado(caso, "")
            # Usar solo sp_name como número de radicado
            no_radicado = radicado if radicado else VALOR_DEFECTO_NA
            matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
            matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
            matriculas_display = ", ".join(matriculas) if matriculas else ""
            
            ws.append([
                codigo_asistente,
                codigo_bot,
                usuario_red,
                nombre_estacion,
                id_proceso,
                no_radicado,
                matriculas_display,
                ESTADO_EXITOSO,
                MENSAJE_PROCESADO_CORRECTAMENTE,
                fecha_inicio,
                hora_inicio,
                fecha_fin,
                hora_fin
            ])
        
        # Agregar casos con error (No Exitosos)
        for item in self.casos_error:
            caso = item.get("caso", {})
            radicado = self._obtener_numero_radicado(caso, "")
            # Usar solo sp_name como número de radicado
            no_radicado = radicado if radicado else VALOR_DEFECTO_NA
            matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
            matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
            matriculas_display = ", ".join(matriculas) if matriculas else ""
            mensaje_error = item.get("mensaje", "")
            
            ws.append([
                codigo_asistente,
                codigo_bot,
                usuario_red,
                nombre_estacion,
                id_proceso,
                no_radicado,
                matriculas_display,
                ESTADO_NO_EXITOSO,
                mensaje_error,
                fecha_inicio,
                hora_inicio,
                fecha_fin,
                hora_fin
            ])
        
        # Agregar casos pendientes
        for item in self.casos_pendientes:
            caso = item.get("caso", {})
            radicado = self._obtener_numero_radicado(caso, "")
            # Usar solo sp_name como número de radicado
            no_radicado = radicado if radicado else VALOR_DEFECTO_NA
            matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
            matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
            matriculas_display = ", ".join(matriculas) if matriculas else ""
            mensaje_pendiente = item.get("mensaje", MENSAJE_NO_PROCESADO)
            
            ws.append([
                codigo_asistente,
                codigo_bot,
                usuario_red,
                nombre_estacion,
                id_proceso,
                no_radicado,
                matriculas_display,
                ESTADO_PENDIENTE,
                mensaje_pendiente,
                fecha_inicio,
                hora_inicio,
                fecha_fin,
                hora_fin
            ])
        
        self._ajustar_ancho_columnas(ws)
        
        ruta_base = self.config.get("Globales", {}).get("RutaBaseProyecto", ".")
        reportes_dir = Path(ruta_base) / RUTA_REPORTES
        reportes_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        reporte_path = reportes_dir / FORMATO_NOMBRE_REPORTE.format(timestamp=timestamp)
        wb.save(str(reporte_path))
        
        # Guardar datos en base de datos
        self._guardar_reporte_en_bd(
            codigo_asistente=codigo_asistente,
            codigo_bot=codigo_bot,
            usuario_red=usuario_red,
            nombre_estacion=nombre_estacion,
            id_proceso=id_proceso,
            fecha_inicio=fecha_inicio,
            hora_inicio=hora_inicio,
            fecha_fin=fecha_fin,
            hora_fin=hora_fin
        )
        
        return str(reporte_path)

    def _enviar_reporte_por_email(self, tipo: str, reporte_path: str, fecha_hora_inicio: datetime, fecha_hora_fin: datetime) -> None:
        """
        Envía el reporte Excel por email al emailResponsable.
        
        Args:
            tipo: Tipo de proceso ("Copias" o "CopiasOficiales")
            reporte_path: Ruta del archivo Excel del reporte
            fecha_hora_inicio: Timestamp de inicio de ejecución
            fecha_hora_fin: Timestamp de fin de ejecución
        """
        try:
            # Obtener emailResponsable según el tipo de proceso
            if tipo == "Copias":
                config_seccion = self.config.get("ReglasNegocio", {}).get("Copias", {})
            else:  # CopiasOficiales
                config_seccion = self.config.get("ReglasNegocio", {}).get("CopiasOficiales", {})
            
            email_responsable = config_seccion.get("emailResponsable", "")
            
            if not email_responsable:
                self.logger.warning(MSG_EMAIL_RESPONSABLE_NO_CONFIG.format(tipo=tipo, accion="reporte por email"))
                return
            
            # Obtener plantilla de email desde config
            reportes_config = self.config.get("Reportes", {})
            plantilla_config = reportes_config.get("PlantillaEmail", {})
            
            if not plantilla_config:
                self.logger.warning(MSG_PLANTILLA_EMAIL_NO_CONFIG)
                return
            
            asunto = plantilla_config.get("asunto", PLANTILLA_REPORTE_ASUNTO)
            cuerpo = plantilla_config.get("cuerpo", PLANTILLA_REPORTE_CUERPO)
            
            # Reemplazar placeholders en la plantilla
            tipo_proceso_display = TIPO_PARTICULARES if tipo == "Copias" else TIPO_OFICIALES
            fecha_reporte = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            fecha_fin = fecha_hora_fin.strftime("%d%m%Y-%H%M%S")
            casos_exitosos = len(self.casos_procesados)
            casos_error = len(self.casos_error)
            casos_pendientes = len(self.casos_pendientes)
            
            asunto = asunto.replace("{tipo_proceso}", tipo_proceso_display)
            cuerpo = cuerpo.replace("{tipo_proceso}", tipo_proceso_display)
            cuerpo = cuerpo.replace("{fecha_reporte}", fecha_reporte)
            cuerpo = cuerpo.replace("{fecha_fin}", fecha_fin)
            cuerpo = cuerpo.replace("{casos_exitosos}", str(casos_exitosos))
            cuerpo = cuerpo.replace("{casos_error}", str(casos_error))
            cuerpo = cuerpo.replace("{casos_pendientes}", str(casos_pendientes))
            
            # Agregar firma al cuerpo
            cuerpo_con_firma = self._agregar_firma(cuerpo)
            
            # Obtener destinatarios según modo QA/PROD
            destinatarios = self._obtener_destinatarios_por_modo(email_responsable)
            
            # Enviar email con adjunto
            self.graph_client.enviar_email(
                usuario_id=self.config.get("GraphAPI", {}).get("user_email", ""),
                asunto=asunto,
                cuerpo=cuerpo_con_firma,
                destinatarios=destinatarios,
                adjuntos=[reporte_path]
            )
            
            self.logger.info(f"Reporte enviado exitosamente por email a {', '.join(destinatarios)} para {tipo}")
        except Exception as e:
            self.logger.error(f"Error enviando reporte por email para {tipo}: {e}")

    def _ajustar_ancho_columnas(self, ws: Any) -> None:
        """
        Ajusta el ancho de las columnas del worksheet según su contenido.

        Args:
            ws: Worksheet de openpyxl a ajustar
        """
        for col in ws.columns:
            max_length = self._calcular_ancho_maximo_columna(col)
            column = col[0].column_letter
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _calcular_ancho_maximo_columna(self, col: Any) -> int:
        """
        Calcula el ancho máximo necesario para una columna.

        Args:
            col: Columna del worksheet

        Returns:
            Ancho máximo encontrado en la columna
        """
        max_length = 0
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        return max_length

    def _guardar_reporte_en_bd(
        self,
        codigo_asistente: str,
        codigo_bot: str,
        usuario_red: str,
        nombre_estacion: str,
        id_proceso: int,
        fecha_inicio: str,
        hora_inicio: str,
        fecha_fin: str,
        hora_fin: str
    ) -> None:
        """
        Guarda los datos del reporte en la base de datos.
        
        Args:
            codigo_asistente: Código del asistente
            codigo_bot: Código del bot
            usuario_red: Usuario de red del bot runner
            nombre_estacion: Nombre de la estación del bot runner
            id_proceso: ID del proceso (PID)
            fecha_inicio: Fecha de inicio (formato YYYY-MM-DD)
            hora_inicio: Hora de inicio (formato HH:MM:SS)
            fecha_fin: Fecha de fin (formato YYYY-MM-DD)
            hora_fin: Hora de fin (formato HH:MM:SS)
        """
        try:
            # Verificar si hay configuración de BD
            db_config = self.config.get("Database", {})
            if not db_config:
                self.logger.warning(MSG_DATABASE_NO_CONFIG)
                return
            
            # Verificar si hay password (debe venir desde rocketbot)
            if not db_config.get("password"):
                self.logger.warning(MSG_DATABASE_PASSWORD_NO_CONFIG)
                return
            
            # Crear instancia de ExpedicionCopiasDB
            db_service = ExpedicionCopiasDB(self.config)
            
            registros_insertados = 0
            registros_fallidos = 0
            
            # Insertar casos exitosos
            for item in self.casos_procesados:
                caso = item.get("caso", {})
                radicado = self._obtener_numero_radicado(caso, "")
                # Usar solo sp_name como número de radicado
                no_radicado = radicado if radicado else VALOR_DEFECTO_NA
                matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
                matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
                matriculas_display = ", ".join(matriculas) if matriculas else ""
                
                if db_service.insert_reporte_expedicion(
                    codigo_asistente=codigo_asistente,
                    codigo_bot=codigo_bot,
                    usuario_red_bot_runner=usuario_red,
                    nombre_estacion_bot_runner=nombre_estacion,
                    id_proceso=id_proceso,
                    no_radicado=no_radicado,
                    matriculas=matriculas_display,
                    estado_proceso=ESTADO_EXITOSO,
                    observacion=MENSAJE_PROCESADO_CORRECTAMENTE,
                    fecha_inicio_ejecucion=fecha_inicio,
                    hora_inicio_ejecucion=hora_inicio,
                    fecha_fin_ejecucion=fecha_fin,
                    hora_fin_ejecucion=hora_fin
                ):
                    registros_insertados += 1
                else:
                    registros_fallidos += 1
            
            # Insertar casos con error (No Exitosos)
            for item in self.casos_error:
                caso = item.get("caso", {})
                radicado = self._obtener_numero_radicado(caso, "")
                # Usar solo sp_name como número de radicado
                no_radicado = radicado if radicado else VALOR_DEFECTO_NA
                matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
                matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
                matriculas_display = ", ".join(matriculas) if matriculas else ""
                mensaje_error = item.get("mensaje", "")
                
                if db_service.insert_reporte_expedicion(
                    codigo_asistente=codigo_asistente,
                    codigo_bot=codigo_bot,
                    usuario_red_bot_runner=usuario_red,
                    nombre_estacion_bot_runner=nombre_estacion,
                    id_proceso=id_proceso,
                    no_radicado=no_radicado,
                    matriculas=matriculas_display,
                    estado_proceso="No Exitoso",
                    observacion=mensaje_error,
                    fecha_inicio_ejecucion=fecha_inicio,
                    hora_inicio_ejecucion=hora_inicio,
                    fecha_fin_ejecucion=fecha_fin,
                    hora_fin_ejecucion=hora_fin
                ):
                    registros_insertados += 1
                else:
                    registros_fallidos += 1
            
            # Insertar casos pendientes
            for item in self.casos_pendientes:
                caso = item.get("caso", {})
                radicado = self._obtener_numero_radicado(caso, "")
                # Usar solo sp_name como número de radicado
                no_radicado = radicado if radicado else VALOR_DEFECTO_NA
                matriculas_str = caso.get("invt_matriculasrequeridas", "") or ""
                matriculas = [m.strip() for m in matriculas_str.split(",") if m.strip()]
                matriculas_display = ", ".join(matriculas) if matriculas else ""
                mensaje_pendiente = item.get("mensaje", MENSAJE_NO_PROCESADO)
                
                if db_service.insert_reporte_expedicion(
                    codigo_asistente=codigo_asistente,
                    codigo_bot=codigo_bot,
                    usuario_red_bot_runner=usuario_red,
                    nombre_estacion_bot_runner=nombre_estacion,
                    id_proceso=id_proceso,
                    no_radicado=no_radicado,
                    matriculas=matriculas_display,
                    estado_proceso=ESTADO_PENDIENTE,
                    observacion=mensaje_pendiente,
                    fecha_inicio_ejecucion=fecha_inicio,
                    hora_inicio_ejecucion=hora_inicio,
                    fecha_fin_ejecucion=fecha_fin,
                    hora_fin_ejecucion=hora_fin
                ):
                    registros_insertados += 1
                else:
                    registros_fallidos += 1
            
            # Cerrar conexión
            db_service.close()
            
            if registros_insertados > 0:
                self.logger.info(f"Reporte guardado en BD: {registros_insertados} registros insertados, {registros_fallidos} fallidos")
            else:
                self.logger.warning(f"No se insertaron registros en BD. Fallidos: {registros_fallidos}")
                
        except Exception as e:
            # No afectar la generación del Excel si falla el guardado en BD
            self.logger.error(f"Error guardando reporte en BD: {e}. El reporte Excel se generó correctamente.")
